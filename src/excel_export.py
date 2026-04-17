"""
excel_export.py — Dashboard'dan Excel indirme.
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime

T = {
    "BG": "1A252F", "FG": "FFFFFF",
    "H1": "1A3A5C", "GRN": "1A7A3E", "RED": "C0392B",
    "GRN_L": "D5F5E3", "RED_L": "FADBD8", "STR": "F4F6F7",
}

def _brd():
    s = Side(style="thin", color="D5D8DC")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, bg, size=10):
    cell.font = Font(name="Arial", bold=True, color=T["FG"], size=size)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _brd()

def _title(ws, text, ncols, bg=None):
    col = get_column_letter(ncols)
    ws.merge_cells(f"A1:{col}1")
    c = ws["A1"]
    c.value = text
    c.font = Font(name="Arial", bold=True, size=13, color=T["FG"])
    c.fill = PatternFill("solid", start_color=bg or T["BG"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

def excel_indir(ana_df: pd.DataFrame, pos_df, baslik="", donem="") -> BytesIO:
    wb = openpyxl.Workbook()
    tarih = datetime.now().strftime("%d.%m.%Y %H:%M")

    # Ana tablo
    ws = wb.active
    ws.title = "ANA TABLO"
    cols = list(ana_df.columns)
    _title(ws, f"{baslik} | {donem} | {tarih}", len(cols), T["H1"])
    for c, h in enumerate(cols, 1):
        _hdr(ws.cell(row=2, column=c, value=h), T["H1"])
    ws.row_dimensions[2].height = 30

    for i, (_, row) in enumerate(ana_df.iterrows()):
        r = i + 3
        bg = T["STR"] if i % 2 else "FFFFFF"
        for c, col in enumerate(cols, 1):
            v = row[col]
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = _brd()
            cell.fill = PatternFill("solid", start_color=bg)
            if isinstance(v, (int, float)):
                if "Fark" in col or "Net" in col or "Yab" in col:
                    cell.number_format = "#,##0;[Red]-#,##0"
                    color = T["GRN"] if v > 0 else T["RED"] if v < 0 else "AAAAAA"
                    cell.font = Font(name="Arial", size=9, color=color)
                elif "pp" in col or "MKK" in col:
                    cell.number_format = "+0.00;-0.00;0.00"
                    color = "1A5276" if v > 0 else T["RED"] if v < 0 else "AAAAAA"
                    cell.font = Font(name="Arial", size=9, color=color)
                elif "%" in col:
                    cell.number_format = "0.00"
                    cell.font = Font(name="Arial", size=9, color="333333")
                else:
                    cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.font = Font(name="Arial", size=10, bold=(c == 1))
                cell.alignment = Alignment(horizontal="left" if c == 1 else "center",
                                           vertical="center")
            ws.row_dimensions[r].height = 17

    ws.column_dimensions["A"].width = 10
    for c in range(2, len(cols)+1):
        ws.column_dimensions[get_column_letter(c)].width = 12
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
