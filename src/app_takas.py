"""
app_takas.py — TAKAS ANALİZİ Sekmesi
app.py'ye import edilir.
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from takas_depo import (
    dosyalar_yukle, donemler_listele, kurumlar_listele,
    kurum_net_pozisyon, alarm_listesi, takas_analiz,
    hisse_kurum_detay, trend_analiz, donem_sil,
    BUYUK_YERLI, AKILLI_PARA, FON_YABANCI, KURUMLAR
)


def aylik_trend_analiz(df_csv: pd.DataFrame) -> dict:
    """6 aylık veriden sinyal tespiti."""
    if df_csv.empty:
        return {}
    df = df_csv[df_csv["tip"] == "aylik"].copy()
    if df.empty:
        return {}
    donemler = sorted(df["donem"].unique())
    if len(donemler) < 2:
        return {}

    sonuclar = {
        "sifirdan_giris": [], "buyuk_alim": [], "duzenli_artis": [],
        "sifira_cikis": [], "buyuk_satis": [], "duzenli_azalis": [],
    }

    for hisse in df["hisse"].unique():
        for kurum in df["kurum"].unique():
            k_df = df[(df["hisse"]==hisse) & (df["kurum"]==kurum)].sort_values("donem")
            if k_df.empty or len(k_df) < 2: continue
            oranlar = k_df["oran2"].tolist()
            d_list  = k_df["donem"].tolist()
            ilk = oranlar[0]; son = oranlar[-1]; fark = son - ilk
            kron = " → ".join([f"%{o:.1f}" for o in oranlar])

            if ilk < 0.1 and son >= 1.0:
                sonuclar["sifirdan_giris"].append({"Hisse":hisse,"Kurum":kurum,
                    "İlk":d_list[0],"Son":d_list[-1],"Son %":round(son,2),"Kronoloji":kron})

            for i in range(1, len(oranlar)):
                d_fark = oranlar[i] - oranlar[i-1]
                if d_fark >= 3.0:
                    sonuclar["buyuk_alim"].append({"Hisse":hisse,"Kurum":kurum,
                        "Dönem":d_list[i],"Önceki %":round(oranlar[i-1],2),
                        "Sonraki %":round(oranlar[i],2),"Artış":round(d_fark,2),"Kronoloji":kron})
                if d_fark <= -3.0:
                    sonuclar["buyuk_satis"].append({"Hisse":hisse,"Kurum":kurum,
                        "Dönem":d_list[i],"Önceki %":round(oranlar[i-1],2),
                        "Sonraki %":round(oranlar[i],2),"Düşüş":round(d_fark,2),"Kronoloji":kron})

            if len(oranlar) >= 3:
                son3 = oranlar[-3:]
                if all(son3[i]>son3[i-1] for i in range(1,len(son3))) and son3[-1]>=0.5:
                    sonuclar["duzenli_artis"].append({"Hisse":hisse,"Kurum":kurum,
                        "Son %":round(son,2),"Toplam":round(fark,2),"Kronoloji":kron})
                if all(son3[i]<son3[i-1] for i in range(1,len(son3))) and son3[0]>=0.5:
                    sonuclar["duzenli_azalis"].append({"Hisse":hisse,"Kurum":kurum,
                        "Son %":round(son,2),"Toplam":round(fark,2),"Kronoloji":kron})

            if ilk >= 1.0 and son < 0.1:
                sonuclar["sifira_cikis"].append({"Hisse":hisse,"Kurum":kurum,
                    "İlk %":round(ilk,2),"Kronoloji":kron})

    return {k: pd.DataFrame(v) for k, v in sonuclar.items()}


def _kart(hisse, kurum, renk, detay, kron):
    st.markdown(
        f"""<div style='border-left:4px solid {renk};padding:5px 10px;
        margin:3px 0;background:#FAFAFA;border-radius:0 4px 4px 0;'>
        <b>{hisse}</b> — <span style='color:{renk}'>{kurum}</span>
        <span style='float:right;font-weight:bold;color:{renk}'>{detay}</span><br>
        <small style='color:#666'>{kron}</small>
        </div>""", unsafe_allow_html=True
    )


def aylik_trend_goster():
    """Aylık trend — hisse bazlı, takip kurumları, %5+ filtresi."""
    from takas_depo import _oku, AKILLI_PARA, KURUMLAR

    df = _oku()
    if df.empty:
        st.info("Aylık veri yok.")
        return

    df_aylik = df[df["tip"] == "aylik"].copy()
    if df_aylik.empty:
        st.info("Aylık veri bulunamadı.")
        return

    donemler = sorted(df_aylik["donem"].unique())
    son_donem = donemler[-1]
    st.caption(f"📅 Mevcut aylar: {' | '.join(donemler)} | **Son: {son_donem}**")

    # ── Filtreler ─────────────────────────────────────────────────────────────
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        sec_ay = st.selectbox("Ay seç", donemler, index=len(donemler)-1, key="trend_ay")
    with col2:
        min_pct = st.number_input("Min pozisyon %", value=5.0, step=1.0, key="trend_min_pct")
    with col3:
        min_kurum = st.number_input("Min kurum sayısı", value=1, step=1, min_value=1, key="trend_min_kurum")

    # ── Seçili ay verisi ──────────────────────────────────────────────────────
    df_sec = df_aylik[df_aylik["donem"] == sec_ay].copy()

    # Sadece takip kurumları
    df_sec = df_sec[df_sec["kurum"].isin(KURUMLAR)]

    # Min %5 filtresi
    df_sec = df_sec[df_sec["oran2"] >= min_pct]

    if df_sec.empty:
        st.warning(f"'{sec_ay}' döneminde %{min_pct}+ pozisyonlu takip kurumu bulunamadı.")
        return

    # ── Önceki ay için trend hesapla ──────────────────────────────────────────
    onceki_ay = donemler[donemler.index(sec_ay)-1] if donemler.index(sec_ay) > 0 else None

    if onceki_ay:
        df_onc = df_aylik[df_aylik["donem"] == onceki_ay][["hisse","kurum","oran2"]].rename(
            columns={"oran2": "onceki_oran"}
        )
        df_sec = df_sec.merge(df_onc, on=["hisse","kurum"], how="left")
        df_sec["degisim"] = (df_sec["oran2"] - df_sec["onceki_oran"].fillna(0)).round(2)
        df_sec["yeni_giris"] = df_sec["onceki_oran"].isna() | (df_sec["onceki_oran"] < 1.0)
    else:
        df_sec["onceki_oran"] = 0
        df_sec["degisim"] = df_sec["oran2"]
        df_sec["yeni_giris"] = True

    # ── Hisse bazlı grupla ────────────────────────────────────────────────────
    rows = []
    for hisse, grp in df_sec.groupby("hisse"):
        grp = grp.sort_values("oran2", ascending=False)

        kurumlar_str = "  ".join([
            f"**{r['kurum']}** %{r['oran2']:.1f}" +
            (f" 🆕" if r["yeni_giris"] else
             f" ▲%{r['degisim']:.1f}" if r["degisim"] > 0.5 else
             f" ▼%{abs(r['degisim']):.1f}" if r["degisim"] < -0.5 else "")
            for _, r in grp.iterrows()
        ])

        toplam    = grp["oran2"].sum()
        kurum_say = len(grp)
        yeni_say  = grp["yeni_giris"].sum()
        artan_say = (grp["degisim"] > 0.5).sum()

        # Alarm seviyesi
        if kurum_say >= 3 and toplam >= 20:
            alarm = "🔴 KRİTİK"
        elif kurum_say >= 2 and toplam >= 10:
            alarm = "🟠 GÜÇLÜ"
        elif yeni_say > 0:
            alarm = "🟡 YENİ GİRİŞ"
        elif artan_say >= 1:
            alarm = "🟢 ARTIYOR"
        else:
            alarm = "⚪ İZLE"

        # Trend ikonu (son 3 ay)
        trend_str = ""
        if onceki_ay:
            trend_vals = []
            for d in donemler[-3:]:
                v = df_aylik[(df_aylik["hisse"]==hisse) & 
                             (df_aylik["donem"]==d) &
                             (df_aylik["kurum"].isin(KURUMLAR))]["oran2"].sum()
                trend_vals.append(v)
            if len(trend_vals) >= 2:
                if all(trend_vals[i] > trend_vals[i-1] for i in range(1, len(trend_vals))):
                    trend_str = "🚀 Sürekli Artış"
                elif trend_vals[-1] > trend_vals[-2]:
                    trend_str = "📈 Son Ay Arttı"
                elif trend_vals[-1] < trend_vals[-2]:
                    trend_str = "📉 Son Ay Azaldı"
                else:
                    trend_str = "➡️ Sabit"

        rows.append({
            "Hisse":        hisse,
            "Kurumlar":     kurumlar_str,
            "Toplam %":     round(toplam, 1),
            "Kurum Sayısı": kurum_say,
            "Trend":        trend_str,
            "Alarm":        alarm,
        })

    df_tablo = pd.DataFrame(rows)
    df_tablo = df_tablo[df_tablo["Kurum Sayısı"] >= int(min_kurum)]
    df_tablo = df_tablo.sort_values(["Alarm","Toplam %"], ascending=[True, False])

    if df_tablo.empty:
        st.warning("Filtre kriterlerine uyan hisse yok.")
        return

    # ── Özet metrikler ────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Toplam Hisse", len(df_tablo))
    c2.metric("🔴 Kritik",    len(df_tablo[df_tablo["Alarm"]=="🔴 KRİTİK"]))
    c3.metric("🟠 Güçlü",     len(df_tablo[df_tablo["Alarm"]=="🟠 GÜÇLÜ"]))
    c4.metric("🟡 Yeni Giriş",len(df_tablo[df_tablo["Alarm"]=="🟡 YENİ GİRİŞ"]))

    st.divider()

    # ── Alarm kartları ────────────────────────────────────────────────────────
    for _, r in df_tablo.iterrows():
        alarm = r["Alarm"]
        renk = {
            "🔴 KRİTİK":    "#C0392B",
            "🟠 GÜÇLÜ":     "#E67E22",
            "🟡 YENİ GİRİŞ":"#F39C12",
            "🟢 ARTIYOR":   "#1A7A3E",
            "⚪ İZLE":       "#717D7E",
        }.get(alarm, "#717D7E")

        st.markdown(
            f"""<div style='border-left:5px solid {renk};padding:8px 12px;
            margin:5px 0;background:#FAFAFA;border-radius:0 6px 6px 0;'>
            <span style='font-size:16px;font-weight:bold;'>{r['Hisse']}</span>
            <span style='color:{renk};margin-left:10px;font-weight:bold;'>{alarm}</span>
            <span style='float:right;color:#888;'>Toplam: <b>%{r['Toplam %']}</b> | {r['Kurum Sayısı']} kurum | {r['Trend']}</span><br>
            <small style='color:#555;'>{r['Kurumlar']}</small>
            </div>""",
            unsafe_allow_html=True
        )


def takas_sekmesi():
    """Ana takas analizi sekmesi."""

    tab1, tab2 = st.tabs(["📊 Takas Analizi", "🧠 Smart Money"])

    with tab1:
        _takas_analiz_tab()

    with tab2:
        _smart_money_tab()


def _smart_money_tab():
    """Smart Money Dashboard — kurum_takas.csv'den direkt hesaplar."""
    import os

    BUYUK_YERLI = ['IS_YATIRIM','HALK_YATIRIM','GARANTI','VAKIF',
                   'AK_YATIRIM','YAPI_KREDI','TEB','ZIRAAT_YATIRIM','DENIZ_YATIRIM']
    FONLAR      = ['YABANCI','YAT_FONLARI','EMEKLILIK','MIDAS']
    AKILLI      = ['MARBAS','BULLS','PUSULA','ALNUS','A1_CAPITAL']
    DAGITICI    = ['INFO','TERA']
    TUM_KURUMLAR = BUYUK_YERLI + FONLAR + AKILLI + DAGITICI

    SENARYOLAR = {
        "🏦 Büyük Yerli → Fon/Yabancı" : ("Büyük Yerli çıkıyor, Fon/Yabancı alıyor — AL sinyali", BUYUK_YERLI, FONLAR),
        "⚡ Dağıtıcı → Büyük Yerli"     : ("INFO/TERA satıyor, Büyük Yerli alıyor — AL sinyali",   DAGITICI,    BUYUK_YERLI),
        "🏦 Büyük Yerli → Akıllı Para"  : ("Büyük Yerli çıkıyor, Akıllı Para alıyor — AL sinyali", BUYUK_YERLI, AKILLI),
        "🔄 Akıllı Para → Dağıtıcı"    : ("MARBAS/BULLS satıyor, INFO/TERA alıyor — DİKKAT",       AKILLI,      DAGITICI),
    }

    MIN_POZISYON     = 1.5
    MIN_HANDOFF_GUCU = 2.0

    def ds(d):
        d = str(d).strip()
        try:
            if len(d)==8 and d.isdigit(): return int(d)
            elif len(d)==7 and '_' in d:
                y,m=d.split('_'); return int(y)*10000+int(m)*100+50
            elif len(d)==9 and '_' in d:
                ym,w=d.split('_'); return int(ym)*100+int(w)*10
        except: pass
        return 0

    @st.cache_data(ttl=1800)
    def handoff_hesapla():
        import os
        # Parquet varsa hızlı oku
        for path in ["src/data/takas/handoff_sonuclar.parquet",
                     "data/takas/handoff_sonuclar.parquet"]:
            if os.path.exists(path):
                return pd.read_parquet(path)
        return pd.DataFrame()

    # ── UI ────────────────────────────────────────────────────────
    st.markdown("### 🧠 Smart Money Dashboard")
    st.caption("kurum_takas.csv'den otomatik hesaplanır · 30dk önbellek")

    col_g, col_b = st.columns([2,5])
    with col_g:
        if st.button("🔄 Yenile", key="sm_yenile"):
            st.cache_data.clear()
            st.rerun()
    with col_b:
        st.info("Yeni takas yükledikten sonra 🔄 Yenile'ye bas.")

    with st.spinner("Handoff hesaplanıyor..."):
        hdf = handoff_hesapla()

    if hdf.empty:
        st.warning("⚠️ Handoff verisi bulunamadı. `bist_app/` klasöründe `python handoff_analiz.py` çalıştırın.")
        return

    st.success(f"✅ {len(hdf):,} handoff | {hdf['hisse'].nunique()} hisse | {hdf['donem'].nunique()} dönem")

    # ── Senaryo Özeti ─────────────────────────────────────────────
    st.markdown("#### 📊 Senaryo Özeti")
    ozet_rows = []
    for senaryo, (aciklama, sat_list, al_list) in SENARYOLAR.items():
        filtre = hdf['satan'].isin(sat_list) & hdf['alan'].isin(al_list)
        df_f   = hdf[filtre]
        son    = df_f[df_f['donem_sira'] >= 20260400]
        vals   = df_f['getiri_60g'].dropna()
        win_60 = round((vals > 0).mean() * 100, 0) if len(vals) >= 3 else None
        ort_60 = round(vals.mean(), 1) if len(vals) >= 3 else None
        ozet_rows.append({
            "Senaryo"         : senaryo,
            "Toplam"          : len(df_f),
            "Win 60g%"        : win_60,
            "Ort 60g%"        : ort_60,
            "Son Dönem Sinyal": len(son),
            "Son Dönem Hisse" : son['hisse'].nunique(),
        })

    ozet_df = pd.DataFrame(ozet_rows)

    def renk_win(val):
        try:
            v = float(val)
            if v >= 65: return "background-color:#1a472a;color:white"
            if v >= 55: return "background-color:#2d6a4f;color:white"
            if v >= 45: return "background-color:#52b788"
        except: pass
        return ""

    st.dataframe(
        ozet_df.style.map(renk_win, subset=["Win 60g%"]),
        use_container_width=True, hide_index=True
    )

    st.divider()

    # ── Aktif Sinyaller ───────────────────────────────────────────
    st.markdown("#### 🔥 Aktif Sinyaller (Son Dönem)")
    senaryo_sec = st.selectbox("Senaryo seç:",
        ["🔍 Tüm Senaryolar"] + list(SENARYOLAR.keys()),
        key="sm_senaryo_sec"
    )

    if senaryo_sec == "🔍 Tüm Senaryolar":
        aktif = hdf[hdf['donem_sira'] >= 20260400].sort_values(
            ['donem_sira','toplam_guc'], ascending=[False,False])
    else:
        sat_list = SENARYOLAR[senaryo_sec][1]
        al_list  = SENARYOLAR[senaryo_sec][2]
        aktif = hdf[
            hdf['satan'].isin(sat_list) &
            hdf['alan'].isin(al_list) &
            (hdf['donem_sira'] >= 20260400)
        ].sort_values(['donem_sira','toplam_guc'], ascending=[False,False])

    hisse_ozet = aktif.groupby('hisse').agg(
        sinyal        = ('hisse','count'),
        max_guc       = ('toplam_guc','max'),
        ilk_donem     = ('donem', 'last'),
        son_donem     = ('donem','first'),
        son_donem_sira= ('donem_sira','max'),
        ilk_donem_sira= ('donem_sira','min'),
        satan_kurumlar= ('satan', lambda x: ', '.join(x.unique()[:3])),
        alan_kurumlar = ('alan',  lambda x: ', '.join(x.unique()[:3])),
    ).reset_index().sort_values('max_guc', ascending=False)

    if hisse_ozet.empty:
        st.info("Bu senaryoda son dönemde sinyal yok.")
        return

    # Arama kutusu
    arama = st.text_input("🔍 Hisse Ara", placeholder="ONCSM, TATEN, ENSRI...",
                          key="sm_arama")
    if arama:
        hisse_ozet = hisse_ozet[
            hisse_ozet['hisse'].str.contains(arama.strip().upper())
        ].reset_index(drop=True)

    st.success(f"✅ **{len(hisse_ozet)} hisse** — {senaryo_sec}")

    # ── Sinyal Hafızası ───────────────────────────────────────────
    import json, os
    from datetime import date, datetime
    HAFIZA_PATH = "src/data/takas/sinyal_hafizasi.json"
    os.makedirs("src/data/takas", exist_ok=True)

    def hafiza_yukle():
        if os.path.exists(HAFIZA_PATH):
            with open(HAFIZA_PATH, encoding='utf-8') as f:
                return json.load(f)
        return {}

    def hafiza_kaydet(h):
        with open(HAFIZA_PATH, 'w', encoding='utf-8') as f:
            json.dump(h, f, ensure_ascii=False, indent=2)

    hafiza = hafiza_yukle()

    # ── Tile Görünümü ─────────────────────────────────────────────
    cols_per_row = 4
    for row_start in range(0, min(len(hisse_ozet), 24), cols_per_row):
        cols = st.columns(cols_per_row)
        for j, col in enumerate(cols):
            idx = row_start + j
            if idx >= len(hisse_ozet): break
            r = hisse_ozet.iloc[idx]
            hisse = r['hisse']
            h_kayit = hafiza.get(hisse, {})

            # Durum rengi
            if r['son_donem_sira'] >= 20260510:
                renk = "#1a472a"  # yeşil — aktif
            elif r['son_donem_sira'] >= 20260400:
                renk = "#7d4e00"  # sarı — zayıflıyor
            else:
                renk = "#4a1a1a"  # kırmızı — eski

            with col:
                if st.button(
                    f"🔍 {hisse}",
                    key=f"tile_{hisse}_{senaryo_sec[:5]}",
                    use_container_width=True
                ):
                    st.session_state['sm_secili_hisse'] = hisse

                st.markdown(
                    f"""<div style='background:{renk};border-radius:6px;
                    padding:8px;margin:-8px 0 8px 0;text-align:center;'>
                    <span style='color:#aaa;font-size:11px;'>Güç: {r['max_guc']:.1f} · {r['sinyal']} sinyal</span><br>
                    <span style='color:#6c9;font-size:10px;'>{r['satan_kurumlar']}</span><br>
                    <span style='color:#888;font-size:10px;'>{r['son_donem']}</span>
                    </div>""",
                    unsafe_allow_html=True
                )

    # ── Seçili Hisse Detayı ───────────────────────────────────────
    secili = st.session_state.get('sm_secili_hisse')
    if secili and secili in hisse_ozet['hisse'].values:
        st.divider()
        st.markdown(f"## 🟢 {secili} — Sinyal Detayı")

        r = hisse_ozet[hisse_ozet['hisse']==secili].iloc[0]

        # ── Dinamik tarih eşikleri ────────────────────────────────
        from datetime import date, timedelta
        bugun_sira = int(date.today().strftime('%Y%m%d'))
        bir_ay_once = int((date.today() - timedelta(days=30)).strftime('%Y%m%d'))
        iki_ay_once = int((date.today() - timedelta(days=60)).strftime('%Y%m%d'))

        # Durum hesapla — dinamik
        if r['son_donem_sira'] >= bir_ay_once:
            durum = "🟢 Aktif"
            durum_renk = "green"
        elif r['son_donem_sira'] >= iki_ay_once:
            durum = "🟡 Zayıflıyor"
            durum_renk = "orange"
        else:
            durum = "🔴 Eski Sinyal"
            durum_renk = "red"

        # ── Güncel fiyat çek ─────────────────────────────────────
        import yfinance as yf
        guncel_fiyat = None
        try:
            fiyat_df = yf.download(f"{secili}.IS", period="10d",
                                   progress=False, auto_adjust=True)
            if isinstance(fiyat_df.columns, pd.MultiIndex):
                fiyat_df.columns = fiyat_df.columns.get_level_values(0)
            if not fiyat_df.empty:
                guncel_fiyat = float(fiyat_df['Close'].dropna().iloc[-1])
        except:
            pass

        # ── Hafıza güncelle ───────────────────────────────────────
        if secili not in hafiza:
            hafiza[secili] = {}

        if 'ilk_sinyal' not in hafiza[secili]:
            hafiza[secili]['ilk_sinyal'] = r['ilk_donem']
        if 'kayit_tarihi' not in hafiza[secili]:
            hafiza[secili]['kayit_tarihi'] = str(date.today())
        if not hafiza[secili].get('giris_fiyat') and guncel_fiyat:
            hafiza[secili]['giris_fiyat'] = guncel_fiyat

        hafiza[secili]['son_guncelleme'] = str(date.today())
        hafiza[secili]['durum'] = durum
        hafiza[secili]['son_guc'] = float(r['max_guc'])
        hafiza[secili]['sinyal_sayisi'] = int(r['sinyal'])
        hafiza_kaydet(hafiza)

        giris_fiyat = hafiza[secili].get('giris_fiyat')
        kayit_tarihi = hafiza[secili].get('kayit_tarihi', str(date.today()))
        try:
            takipte_gun = (date.today() - date.fromisoformat(kayit_tarihi)).days
        except:
            takipte_gun = 0

        if guncel_fiyat and giris_fiyat and float(giris_fiyat) > 0:
            getiri = round((guncel_fiyat / float(giris_fiyat) - 1) * 100, 2)
        else:
            getiri = None

        # ── Metrik satır 1 ────────────────────────────────────────
        c1, c2, c3 = st.columns(3)
        c1.metric("📅 İlk Sinyal", r['ilk_donem'])
        c2.metric("🔢 Sinyal Adedi", r['sinyal'])
        c3.metric("💪 Max Güç", f"{r['max_guc']:.1f}")

        # ── Metrik satır 2 ────────────────────────────────────────
        c4, c5, c6, c7 = st.columns(4)
        c4.metric("💰 Giriş Fiyatı", f"{float(giris_fiyat):.2f} ₺" if giris_fiyat else "—")
        c5.metric("📈 Güncel Fiyat", f"{guncel_fiyat:.2f} ₺" if guncel_fiyat else "—")
        if getiri is not None:
            c6.metric("🎯 Getiri", f"%{getiri:+.1f}", delta=f"{getiri:+.1f}%")
        else:
            c6.metric("🎯 Getiri", "—")
        c7.metric("⏱️ Takipte", f"{takipte_gun} gün", delta=durum)

        # ── Satan / Alan ──────────────────────────────────────────
        c8, c9 = st.columns(2)
        c8.metric("📤 Satan", r['satan_kurumlar'])
        c9.metric("📥 Alan", r['alan_kurumlar'])

        # ── Not alanı — otomatik kayıt ────────────────────────────
        mevcut_not = hafiza[secili].get('not', '')
        yeni_not = st.text_area("📝 Not", value=mevcut_not,
                                key=f"not_{secili}", height=80,
                                placeholder="Örn: Fon alımı devam ediyor, ENSRI MSCI beklentisi...")
        if yeni_not != mevcut_not:
            hafiza[secili]['not'] = yeni_not
            hafiza_kaydet(hafiza)

        # ── Kronoloji ─────────────────────────────────────────────
        st.markdown("#### 📜 Handoff Kronolojisi")
        gecmis = hdf[hdf['hisse']==secili].sort_values('donem_sira', ascending=False)

        def renk_satir(row):
            alan  = str(row.get('alan',''))
            satan = str(row.get('satan',''))
            if any(k in alan  for k in ['YABANCI','YAT_FONLARI','EMEKLILIK','MIDAS']):
                return ['background-color:#d5f5e3;color:#1a5276']*len(row)
            if any(k in satan for k in ['INFO','TERA']):
                return ['background-color:#fadbd8;color:#922b21']*len(row)
            return ['']*len(row)

        cols_goster = ['donem','alarm_tipi','satan','satan_once%','satan_simdi%',
                       'alan','alan_once%','alan_simdi%']
        # Sadece mevcut kolonları göster
        cols_var = [c for c in cols_goster if c in gecmis.columns]
        if 'toplam_guc' not in cols_var:
            cols_var.append('toplam_guc')
        styled = gecmis[cols_var].head(20).style.apply(renk_satir, axis=1)
        st.dataframe(styled, use_container_width=True, hide_index=True)

        # ── Özet metrikler ────────────────────────────────────────
        cm1, cm2, cm3 = st.columns(3)
        cm1.metric("Toplam Handoff", len(gecmis))
        cm2.metric("Net Akış Gücü", f"{gecmis['toplam_guc'].sum():.1f}")
        cm3.metric("En Yüksek Güç", f"{gecmis['toplam_guc'].max():.1f}")

        # En aktif kurumlar
        ca1, ca2 = st.columns(2)
        ca1.metric("En Aktif Satan",
                   gecmis['satan'].value_counts().index[0] if len(gecmis) > 0 else "—")
        ca2.metric("En Aktif Alan",
                   gecmis['alan'].value_counts().index[0] if len(gecmis) > 0 else "—")

    # ── Tüm Açık Sinyaller Tablosu ────────────────────────────────
    with st.expander("📊 Tüm Açık Sinyaller (Hafıza)", expanded=False):
        hafiza = hafiza_yukle()
        if hafiza:
            satirlar = []
            for hisse, kayit in hafiza.items():
                satirlar.append({
                    'Hisse'       : hisse,
                    'İlk Sinyal'  : kayit.get('ilk_sinyal','—'),
                    'Kayıt'       : kayit.get('kayit_tarihi','—'),
                    'Güç'         : kayit.get('son_guc', 0),
                    'Sinyal'      : kayit.get('sinyal_sayisi', 0),
                    'Durum'       : kayit.get('durum','—'),
                    'Giriş ₺'     : kayit.get('giris_fiyat'),
                    'Not'         : kayit.get('not',''),
                })
            hafiza_df = pd.DataFrame(satirlar).sort_values('Güç', ascending=False)
            st.dataframe(hafiza_df, use_container_width=True, hide_index=True)
        else:
            st.info("Henüz takip edilen sinyal yok. Tile'a tıklayarak ekle.")

    st.divider()
    st.dataframe(hisse_ozet, use_container_width=True, hide_index=True)


# ── CACHE'Lİ HESAPLAMA ────────────────────────────────────────────────────────
@st.cache_data(ttl=600)
def _birikimli_hesapla(tip: str, son_x_donem: int, min_ardisik: int = 2) -> tuple:
    """
    Tüm geçmişten kümülatif net alım/satım hesaplar.
    tip: haftalik / aylik / gunluk
    son_x_donem: son kaç döneme bak
    Returns: (toplayanlar, satanlar, son_donem, donem_listesi)
    """
    from takas_depo import _oku, BUYUK_YERLI, AKILLI_PARA, FON_YABANCI
    df = _oku()
    if df.empty:
        return [], [], None, []

    df = df[df["tip"] == tip].copy()
    if df.empty:
        return [], [], None, []

    # Short kapama düzeltmesi
    if "short_kapama" in df.columns:
        df.loc[df["short_kapama"] == True, "dolasim_pct"] = \
            df.loc[df["short_kapama"] == True, "oran2"]

    donemler = sorted(df["donem"].astype(str).unique())
    if not donemler:
        return [], [], None, []

    son_donem    = donemler[-1]
    secili_donm  = donemler[-son_x_donem:]

    df_sec = df[df["donem"].isin(secili_donm)].copy()

    def grup(k):
        if k in BUYUK_YERLI: return "Büyük Yerli"
        if k in AKILLI_PARA: return "Akıllı Para"
        if k in FON_YABANCI: return "Fon/Yabancı"
        return "Diğer"

    toplayanlar, satanlar = [], []

    for (hisse, kurum), kdf in df_sec.sort_values("donem").groupby(["hisse", "kurum"]):
        if len(kdf) < min_ardisik:
            continue
        vals      = kdf["dolasim_pct"].tolist()
        son_n     = vals  # tüm seçili dönemler
        oran2     = float(kdf["oran2"].iloc[-1])
        trend_str = " → ".join([f"{v:+.1f}%" for v in son_n[-4:]])
        toplam    = round(sum(vals), 2)
        g         = grup(kurum)

        if sum(son_n) >= 3:
            surekli = all(v > 0 for v in son_n) and all(son_n[i] >= son_n[i-1] for i in range(1, len(son_n)))
            toplayanlar.append({
                "hisse": hisse, "kurum": kurum, "grup": g,
                "toplam": toplam, "son_pct": round(son_n[-1], 2),
                "oran2": round(oran2, 2), "donem_say": len(kdf),
                "surekli": surekli, "trend_str": trend_str,
            })
        elif sum(son_n) <= -3:
            surekli = all(v < 0 for v in son_n) and all(son_n[i] <= son_n[i-1] for i in range(1, len(son_n)))
            satanlar.append({
                "hisse": hisse, "kurum": kurum, "grup": g,
                "toplam": abs(toplam), "son_pct": abs(round(son_n[-1], 2)),
                "oran2": round(oran2, 2), "donem_say": len(kdf),
                "surekli": surekli, "trend_str": trend_str,
            })

    return toplayanlar, satanlar, son_donem, secili_donm


@st.cache_data(ttl=300)
def _fiyat_cek(hisse: str) -> tuple:
    """yfinance'dan son kapanış + % değişim. Returns (fiyat, degisim_pct)"""
    try:
        import yfinance as yf
        t = yf.Ticker(f"{hisse}.IS")
        h = t.history(period="5d")
        if len(h) >= 2:
            son   = float(h["Close"].iloc[-1])
            onceki= float(h["Close"].iloc[-2])
            return round(son, 2), round((son/onceki - 1)*100, 2)
        elif len(h) == 1:
            return round(float(h["Close"].iloc[-1]), 2), None
    except:
        pass
    return None, None


def _detay_panel_inline(hisse: str, secili_donm: list = None):
    """
    Seçili hissenin seçilen dönemdeki net alış/satış dağılımı.
    Matriks gibi: sol=alanlar, sağ=satanlar + tahta değişimi özeti.
    """
    from takas_depo import _oku
    df = _oku()
    if df.empty:
        return
    h_df = df[df["hisse"] == hisse].copy()
    if h_df.empty:
        return

    # Short kapama düzeltmesi
    if "short_kapama" in h_df.columns:
        h_df.loc[h_df["short_kapama"] == True, "dolasim_pct"] = \
            h_df.loc[h_df["short_kapama"] == True, "oran2"]

    # Dönem filtresi
    if secili_donm:
        h_df = h_df[h_df["donem"].isin(secili_donm)]

    if h_df.empty:
        st.caption("Seçilen dönemde veri yok.")
        return

    # ── Tahta (tks2) değişimi ─────────────────────────────────────────────────
    # İlk ve son dönem tks2 ortalamasından değişim hesapla
    tks2_ilk = h_df.groupby("donem")["tks2"].mean()
    if len(tks2_ilk) >= 2:
        tks2_bas  = float(tks2_ilk.iloc[0])
        tks2_son  = float(tks2_ilk.iloc[-1])
        tks2_fark = tks2_son - tks2_bas
        tks2_pct  = round((tks2_fark / tks2_bas * 100), 2) if tks2_bas > 0 else 0
    else:
        tks2_fark = 0
        tks2_pct  = 0

    # ── Kurum bazında kümülatif topla ─────────────────────────────────────────
    ozet = h_df.groupby("kurum").agg(
        net_pct  =("dolasim_pct", "sum"),
        net_adet =("adet_fark",   "sum"),
        t2_oran  =("oran2",       "last"),
    ).reset_index()
    ozet["net_pct"]  = ozet["net_pct"].round(2)
    ozet["net_adet"] = ozet["net_adet"].round(0).astype(int)

    alanlar  = ozet[ozet["net_pct"] > 0].sort_values("net_pct", ascending=False)
    satanlar = ozet[ozet["net_pct"] < 0].sort_values("net_pct", ascending=True)

    kurum_alan_top  = round(alanlar["net_pct"].sum(), 2)
    kurum_satan_top = round(satanlar["net_pct"].sum(), 2)
    kurum_net       = round(kurum_alan_top + kurum_satan_top, 2)

    # ── Tahta yorumu ──────────────────────────────────────────────────────────
    if tks2_pct < -1 and kurum_net > 0:
        tahta_yorum = "⚠️ Dağıtım — Kurumlar alıyor ama tahta küçülüyor"
        tahta_renk  = "#C0392B"
    elif tks2_pct > 1 and kurum_net > 0:
        tahta_yorum = "✅ Gerçek Birikim — Tahta büyüyor, kurumlar alıyor"
        tahta_renk  = "#1A7A3E"
    elif tks2_pct < -1 and kurum_net < 0:
        tahta_yorum = "📉 Çıkış — Hem tahta küçülüyor hem kurumlar satıyor"
        tahta_renk  = "#7B0000"
    elif tks2_pct > 1 and kurum_net < 0:
        tahta_yorum = "🔄 Mal Değişimi — Tahta büyüyor, kurumlar satıyor"
        tahta_renk  = "#E67E22"
    else:
        tahta_yorum = "➡️ Sabit"
        tahta_renk  = "#888"

    donem_str = f"{secili_donm[0]} → {secili_donm[-1]}" if secili_donm else "Tüm dönemler"

    # ── Başlık + Tahta özeti ──────────────────────────────────────────────────
    st.markdown(
        f"<div style='background:#EAF4FB;border-left:4px solid #1A5276;"
        f"border-radius:4px;padding:6px 12px;margin:4px 0;'>"
        f"<b>📊 {hisse}</b> "
        f"<span style='font-size:11px;color:#888;'>· {donem_str}</span>"
        f"</div>",
        unsafe_allow_html=True
    )

    # Tahta özet satırı
    tks2_renk = "#C0392B" if tks2_pct < 0 else "#1A7A3E"
    st.markdown(
        f"<div style='background:#F8F9FA;border-radius:4px;padding:6px 12px;"
        f"margin:4px 0;display:flex;justify-content:space-between;align-items:center;'>"
        f"<span style='font-size:12px;'>"
        f"📦 Tahta: <b style='color:{tks2_renk};'>{tks2_pct:+.2f}%</b>"
        f" &nbsp;|&nbsp; "
        f"🏦 Kurumlar net: <b style='color:{'#1A7A3E' if kurum_net>0 else '#C0392B'};'>"
        f"{kurum_net:+.2f}%</b>"
        f" &nbsp;|&nbsp; "
        f"<b style='color:{tahta_renk};'>{tahta_yorum}</b>"
        f"</span></div>",
        unsafe_allow_html=True
    )

    ca, cs = st.columns(2)

    with ca:
        st.markdown("**📈 Net Alanlar**")
        gorulu = alanlar[alanlar["net_pct"] >= 2]
        if gorulu.empty:
            st.caption("—")
        else:
            toplam_alan = gorulu["net_pct"].sum()
            for _, r in gorulu.iterrows():
                pay = round(r["net_pct"] / toplam_alan * 100, 0) if toplam_alan else 0
                st.markdown(
                    f"<div style='border-left:3px solid #1A5276;padding:3px 8px;"
                    f"margin:2px 0;background:#F8FBFF;'>"
                    f"<b style='color:#1A5276;font-size:12px;'>{r['kurum']}</b>"
                    f"<span style='float:right;color:#1A7A3E;font-weight:bold;font-size:12px;'>"
                    f"+{r['net_pct']:.2f}%</span><br>"
                    f"<span style='font-size:10px;color:#888;'>"
                    f"T2:%{r['t2_oran']:.2f} · "
                    f"{r['net_adet']:+,} adet · %{pay:.0f} pay</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )

    with cs:
        st.markdown("**📉 Net Satanlar**")
        gorulu_sat = satanlar[satanlar["net_pct"] <= -2]
        if gorulu_sat.empty:
            st.caption("Eşiği geçen satış yok")
        else:
            toplam_satan = abs(gorulu_sat["net_pct"].sum())
            for _, r in gorulu_sat.iterrows():
                pay = round(abs(r["net_pct"]) / toplam_satan * 100, 0) if toplam_satan else 0
                st.markdown(
                    f"<div style='border-left:3px solid #C0392B;padding:3px 8px;"
                    f"margin:2px 0;background:#FFF8F8;'>"
                    f"<b style='color:#C0392B;font-size:12px;'>{r['kurum']}</b>"
                    f"<span style='float:right;color:#C0392B;font-weight:bold;font-size:12px;'>"
                    f"{r['net_pct']:.2f}%</span><br>"
                    f"<span style='font-size:10px;color:#888;'>"
                    f"T2:%{r['t2_oran']:.2f} · "
                    f"{r['net_adet']:+,} adet · %{pay:.0f} pay</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )


@st.cache_data(ttl=600)
@st.cache_data(ttl=600)
def _altin_oran_hesapla(tip: str = "haftalik", son_x_donem: int = 12,
                         min_kons: float = 0, min_net: float = 0,
                         kons_esik: float = 60) -> list:
    """
    Altın Oran — İlk 5 kurum diğerlerinden baskın mı?
    """
    from takas_depo import _oku
    df = _oku()
    if df.empty:
        return []

    df = df[df["tip"] == tip].copy()
    donemler = sorted(df["donem"].astype(str).unique())
    if len(donemler) < 3:
        return []

    secili_donm = donemler[-son_x_donem:]
    donem_sira  = {d: i for i, d in enumerate(donemler)}

    sonuclar = []

    for hisse, h_df in df.groupby("hisse"):
        altin_kayit = []

        for donem in secili_donm:
            d_df = h_df[h_df["donem"] == donem].copy()
            if d_df.empty:
                continue

            d_df = d_df[d_df["oran2"] > 0].drop_duplicates(subset=["kurum"])
            toplam_oran = d_df["oran2"].sum()
            if toplam_oran < 8:
                continue

            d_sorted    = d_df.sort_values("oran2", ascending=False)
            ilk5_toplam = d_sorted.head(5)["oran2"].sum()
            kons        = round((ilk5_toplam / toplam_oran) * 100, 1)

            if kons >= kons_esik:
                altin_kayit.append({
                    "donem"       : donem,
                    "kons"        : kons,
                    "toplam_oran" : round(toplam_oran, 2),
                    "ilk5"        : d_sorted.head(5)[["kurum","oran2"]].to_dict("records"),
                })

        if not altin_kayit:
            continue

        en_son = max(altin_kayit, key=lambda x: x["donem"])

        if min_kons > 0 and en_son["kons"] < min_kons:
            continue

        sonuclar.append({
            "hisse"          : hisse,
            "son_altin_donem": en_son["donem"],
            "konsantrasyon"  : en_son["kons"],
            "altin_sayisi"   : len(altin_kayit),
            "ilk5"           : en_son["ilk5"],
            "ilk3"           : [k["kurum"] for k in en_son["ilk5"][:3]],
            "ilk3_deg"       : 0.0,
            "diger_deg"      : 0.0,
            "kurumlar"       : [{"kurum": k["kurum"], "oran2": k["oran2"],
                                  "net_degisim": 0.0, "grup": ""} for k in en_son["ilk5"]],
            "kronoloji"      : [],
            "kons_tarihleri" : [f"{k['donem']}(%{k['kons']})" for k in altin_kayit[-3:]],
            "skor"           : donem_sira.get(en_son["donem"], 0),
        })

    return sorted(sonuclar,
                  key=lambda x: (x["son_altin_donem"], x["konsantrasyon"]),
                  reverse=True)


def _birikimli_tab():
    """📈 BİRİKMİŞ TAKİP — Sol: Tahta hakimiyeti | Sağ: Kronolojik büyük girişler."""
    from datetime import date, timedelta
    from takas_depo import _oku as _t2_raw

    # ── Tarih aralığı seçimi ──────────────────────────────────────────────────
    df_tum = _t2_raw()
    bugun = date.today()

    if not df_tum.empty:
        # Tüm dönemlerin tarihlerini bul (günlük + haftalık + aylık)
        tum_tarihler = []
        for d in df_tum["donem"].astype(str).unique():
            t = _donem_tarih_cevir(d)
            if t:
                tum_tarihler.append(t)
        if tum_tarihler:
            ilk_tarih = min(tum_tarihler)
            son_tarih = max(tum_tarihler)
        else:
            son_tarih = bugun
            ilk_tarih = bugun - timedelta(days=90)
    else:
        son_tarih = bugun
        ilk_tarih = bugun - timedelta(days=90)

    c1, c2, c3 = st.columns([1.5, 1.5, 1])
    with c1:
        bas_tarih = st.date_input(
            "📅 Başlangıç:", value=ilk_tarih,
            min_value=ilk_tarih, max_value=son_tarih,
            key="bir_bas_tarih", format="DD.MM.YYYY"
        )
    with c2:
        bit_tarih = st.date_input(
            "📅 Bitiş:", value=son_tarih,
            min_value=ilk_tarih, max_value=son_tarih,
            key="bir_bit_tarih", format="DD.MM.YYYY"
        )
    with c3:
        min_artis = st.number_input(
            "Min Artış %:", value=5.0, step=1.0,
            min_value=1.0, max_value=100.0,
            key="bir_min_artis"
        )

    if bas_tarih > bit_tarih:
        st.error("Başlangıç tarihi bitiş tarihinden büyük olamaz.")
        return

    if df_tum.empty:
        st.info("📂 Henüz veri yok.")
        return

    # ── Tarih aralığına giren dönemler (günlük öncelikli, yoksa haftalık, yoksa aylık) ──
    def _tip_once(tip):
        secili = []
        for d in df_tum[df_tum["tip"] == tip]["donem"].astype(str).unique():
            t = _donem_tarih_cevir(d)
            if t and bas_tarih <= t <= bit_tarih:
                secili.append(d)
        return secili

    secili_donemler = _tip_once("gunluk")
    kullanilan_tip = "gunluk"
    if not secili_donemler:
        secili_donemler = _tip_once("haftalik")
        kullanilan_tip = "haftalik"
    if not secili_donemler:
        secili_donemler = _tip_once("aylik")
        kullanilan_tip = "aylik"

    if not secili_donemler:
        st.warning("Seçilen tarih aralığında veri bulunamadı.")
        return

    secili_donemler_sorted = sorted(secili_donemler)
    son_donem = secili_donemler_sorted[-1]

    st.caption(
        f"📅 **{bas_tarih.strftime('%d.%m.%Y')}** → **{bit_tarih.strftime('%d.%m.%Y')}** "
        f"| {len(secili_donemler)} dönem ({kullanilan_tip}) | Son dönem: **{son_donem}**"
    )

    # ── Veri hazırlık ─────────────────────────────────────────────────────────
    df_sec = df_tum[df_tum["donem"].isin(secili_donemler)].copy()
    df_son = df_tum[df_tum["donem"] == son_donem].copy()

    # ── SOL TARAF: Son günün T2'si ile tahta hakimiyeti ───────────────────────
    # ── SAĞ TARAF: Kronolojik büyük girişler ─────────────────────────────────
    col_sol, col_sag = st.columns([1, 1.5])

    # ── SOL: %50+ tahta hakimiyeti ────────────────────────────────────────────
    with col_sol:
        st.markdown("### 🏆 Tahta Hakimiyeti (Son Gün T2)")
        st.caption("İlk 5 kurumun T2 toplamı %50+ olan hisseler")

        if df_son.empty:
            st.info("Son dönem verisi yok.")
        else:
            # Her hisse için ilk 5 kurumun T2 toplamını hesapla
            hakimiyet_list = []
            for hisse, h_df in df_son.groupby("hisse"):
                h_df = h_df[h_df["oran2"] > 0].drop_duplicates(subset=["kurum"])
                toplam_oran = h_df["oran2"].sum()
                if toplam_oran < 5:
                    continue
                ilk5 = h_df.sort_values("oran2", ascending=False).head(5)
                ilk5_top = ilk5["oran2"].sum()
                kons = round(ilk5_top / toplam_oran * 100, 1) if toplam_oran > 0 else 0
                if kons >= 50:
                    hakimiyet_list.append({
                        "hisse": hisse,
                        "kons": kons,
                        "toplam_oran": round(toplam_oran, 1),
                        "ilk5": ilk5[["kurum", "oran2"]].to_dict("records"),
                    })

            hakimiyet_list = sorted(hakimiyet_list, key=lambda x: -x["kons"])
            st.caption(f"**{len(hakimiyet_list)}** hisse")

            for r in hakimiyet_list[:50]:
                renk = "#C0392B" if r["kons"] >= 80 else "#E67E22" if r["kons"] >= 65 else "#1A5276"
                kurumlar_str = " · ".join([
                    f"{k['kurum']} %{k['oran2']:.1f}"
                    for k in r["ilk5"][:3]
                ])
                with st.expander(
                    f"{r['hisse']}  —  🏆 %{r['kons']:.0f} konsantrasyon",
                    expanded=False
                ):
                    st.markdown(
                        f"<div style='font-size:12px;'>"
                        + "".join([
                            f"<div style='padding:2px 0;'>"
                            f"<b style='color:{renk};'>{k['kurum']}</b>"
                            f"<span style='float:right;'>T2: %{k['oran2']:.2f}</span></div>"
                            for k in r["ilk5"]
                        ])
                        + f"<div style='margin-top:4px;color:#888;font-size:11px;'>"
                        f"Toplam T2: %{r['toplam_oran']:.1f} | İlk5: %{r['kons']:.1f}</div>"
                        + "</div>",
                        unsafe_allow_html=True
                    )

    # ── SAĞ: Kronolojik büyük girişler ───────────────────────────────────────
    with col_sag:
        st.markdown("### 📈 Kronolojik Büyük Girişler")
        st.caption(f"Dönem içinde %{min_artis:.0f}+ oran artışı yapan hareketler")

        # Her hisse+kurum için dönem bazında oran değişimi hesapla
        # Önceki dönem oran2 vs şimdiki dönem oran2
        buyuk_girisler = []

        for (hisse, kurum), kdf in df_sec.sort_values("donem").groupby(["hisse", "kurum"]):
            kdf = kdf.sort_values("donem")
            donemler_list = kdf["donem"].tolist()
            oran2_list = kdf["oran2"].tolist()

            for i in range(1, len(donemler_list)):
                onceki = oran2_list[i-1]
                yeni   = oran2_list[i]
                artis  = round(yeni - onceki, 2)
                if artis >= min_artis:
                    buyuk_girisler.append({
                        "hisse":  hisse,
                        "kurum":  kurum,
                        "donem":  donemler_list[i],
                        "onceki": round(onceki, 2),
                        "yeni":   round(yeni, 2),
                        "artis":  artis,
                    })



        buyuk_girisler = sorted(buyuk_girisler, key=lambda x: -x["artis"])

        st.caption(f"**{len(buyuk_girisler)}** hareket tespit edildi")

        if not buyuk_girisler:
            st.info(f"Min %{min_artis:.0f} eşiğini geçen hareket yok.")
        else:
            # Tablo olarak göster
            import pandas as pd
            df_goster = pd.DataFrame(buyuk_girisler)[["hisse","kurum","donem","onceki","yeni","artis"]]
            df_goster.columns = ["Hisse","Kurum","Dönem","Önceki%","Yeni%","Artış%"]
            df_goster = df_goster.sort_values("Artış%", ascending=False)

            st.dataframe(
                df_goster,
                use_container_width=True,
                hide_index=True,
                height=600,
                column_config={
                    "Artış%": st.column_config.NumberColumn(format="%.2f"),
                    "Önceki%": st.column_config.NumberColumn(format="%.2f"),
                    "Yeni%": st.column_config.NumberColumn(format="%.2f"),
                }
            )


def _donem_tarih_cevir(donem: str):
    """
    Dönem adını (tarih, haftalık, aylık) datetime.date'e çevirir.
    20260420        → 2026-04-20
    202604_01       → 2026-04-07  (1. hafta ≈ ilk Pazartesi)
    2026_04         → 2026-04-01
    2026_04_03      → 2026-04-01  (eski aylık format)
    """
    from datetime import date
    import re
    d = str(donem).strip()
    # Günlük: 20260420
    if re.match(r'^\d{8}$', d):
        try:
            return date(int(d[:4]), int(d[4:6]), int(d[6:8]))
        except:
            return None
    # Haftalık: 202604_01
    m = re.match(r'^(\d{4})(\d{2})_(\d{2})$', d)
    if m:
        yil, ay, hafta = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            # Ayın ilk Pazartesisi + (hafta-1)*7
            from datetime import timedelta
            ilk = date(yil, ay, 1)
            # İlk Pazartesi
            ilk_pzt = ilk + timedelta(days=(7 - ilk.weekday()) % 7)
            return ilk_pzt + timedelta(weeks=hafta - 1)
        except:
            return None
    # Aylık: 2026_04 veya 2026_04_03
    m = re.match(r'^(\d{4})_(\d{2})', d)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), 1)
        except:
            return None
    return None


def _takas_analiz_tab():
    """Takas Analizi sekmesi."""
    from datetime import date, timedelta

    # ── BİRİKMİŞ TAKİP butonu ────────────────────────────────────────────────
    if "takas_mod" not in st.session_state:
        st.session_state["takas_mod"] = "tarih"
    mod = st.session_state["takas_mod"]

    b1, _sp = st.columns([1.5, 6.5])
    with b1:
        if st.button("📈 BİRİKMİŞ TAKİP", use_container_width=True,
                     type="primary" if mod == "birikimli" else "secondary",
                     key="tb_birikimli"):
            if mod == "birikimli":
                st.session_state["takas_mod"] = "tarih"
            else:
                st.session_state["takas_mod"] = "birikimli"
                st.cache_data.clear()
            st.rerun()

    st.markdown("---")

    # ── BİRİKMİŞ TAKİP modu ──────────────────────────────────────────────────
    if mod == "birikimli":
        _birikimli_tab()
        return

    # ── TARİH ARALIĞI SEÇİMİ ─────────────────────────────────────────────────
    from takas_depo import _oku as _t2_raw
    df_tumü = _t2_raw()

    # Mevcut günlük dönemleri bul — tarih aralığı için min/max
    gunluk_donemler = []
    if not df_tumü.empty:
        gd = df_tumü[df_tumü["tip"] == "gunluk"]["donem"].astype(str).unique()
        gunluk_donemler = sorted(gd.tolist())

    # Varsayılan tarih aralığı: son 5 iş günü
    bugun = date.today()
    if gunluk_donemler:
        try:
            son_tarih = _donem_tarih_cevir(gunluk_donemler[-1]) or bugun
            ilk_tarih = _donem_tarih_cevir(gunluk_donemler[0]) or (bugun - timedelta(days=30))
        except:
            son_tarih = bugun
            ilk_tarih = bugun - timedelta(days=30)
    else:
        son_tarih = bugun
        ilk_tarih = bugun - timedelta(days=7)

    c1, c2, c3, c4 = st.columns([1.5, 1.5, 1, 1])
    with c1:
        bas_tarih = st.date_input(
            "📅 Başlangıç:", value=son_tarih - timedelta(days=6),
            min_value=ilk_tarih, max_value=son_tarih,
            key="takas_bas_tarih", format="DD.MM.YYYY"
        )
    with c2:
        bit_tarih = st.date_input(
            "📅 Bitiş:", value=son_tarih,
            min_value=ilk_tarih, max_value=son_tarih,
            key="takas_bit_tarih", format="DD.MM.YYYY"
        )
    with c3:
        min_pct = st.number_input("Min %:", value=0.5, step=0.1,
                                   min_value=0.0, max_value=10.0,
                                   key="takas_min_pct")
    with c4:
        st.markdown("<br>", unsafe_allow_html=True)
        ilk_giris = st.checkbox("🔵 İlk Giriş (%0→%3)", key="takas_ilk_giris")

    if bas_tarih > bit_tarih:
        st.error("Başlangıç tarihi bitiş tarihinden büyük olamaz.")
        return

    # ── Seçilen tarih aralığına giren dönemleri bul ───────────────────────────
    if df_tumü.empty:
        st.info("📂 Henüz veri yok. **Veri Yükle** sekmesinden ekleyin.")
        return

    tum_donemler = df_tumü["donem"].astype(str).unique().tolist()
    secili_donemler = []
    for d in tum_donemler:
        t = _donem_tarih_cevir(d)
        if t and bas_tarih <= t <= bit_tarih:
            secili_donemler.append(d)

    if not secili_donemler:
        st.warning(f"⚠️ {bas_tarih.strftime('%d.%m.%Y')} — {bit_tarih.strftime('%d.%m.%Y')} aralığında veri bulunamadı.")
        return

    # Kullanıcıya özet bilgi
    st.caption(
        f"📊 **{bas_tarih.strftime('%d.%m.%Y')}** → **{bit_tarih.strftime('%d.%m.%Y')}** "
        f"| {len(secili_donemler)} dönem: {', '.join(sorted(secili_donemler))}"
    )

    # mod değişkeni aşağıda takas_analiz() için gerekli — günlük olarak geçir
    mod = "gunluk"

    # ── %3+ Artış Yapan Kurumlar ──────────────────────────────────────────────
    from takas_depo import _oku as _t2
    t2_df = _t2()

    if not t2_df.empty:
        t2_sec = t2_df[t2_df["donem"].isin(secili_donemler)].copy()

        if "short_kapama" in t2_sec.columns:
            short_df     = t2_sec[t2_sec["short_kapama"] == True].copy()
            t2_sec_temiz = t2_sec[t2_sec["short_kapama"] != True].copy()
        else:
            short_df     = pd.DataFrame()
            t2_sec_temiz = t2_sec.copy()

        artis_df = t2_sec_temiz.groupby(["hisse","kurum"]).agg(
            dolasim_pct=("dolasim_pct","sum"),
            oran2=("oran2","last")
        ).reset_index()

        if not short_df.empty:
            s_ozet = short_df.groupby(["kurum","hisse"]).agg(
                oran2=("oran2","last"), dolasim_pct=("dolasim_pct","sum")
            ).reset_index()
            s_ozet = s_ozet[s_ozet["dolasim_pct"] >= max(3, min_pct)]
            if not s_ozet.empty:
                with st.expander("⚠️ Short Kapama Tespit Edildi", expanded=False):
                    for _, r in s_ozet.iterrows():
                        st.markdown(
                            f"<span style='font-size:12px;color:#E67E22;'>"
                            f"<b>{r['kurum']}</b> → <b>{r['hisse']}</b> "
                            f"T2:%{r['oran2']:.2f} Giriş:%{r['dolasim_pct']:.1f}</span>",
                            unsafe_allow_html=True
                        )

        if ilk_giris:
            df_onc = t2_df[~t2_df["donem"].isin(secili_donemler)].copy()
            onc    = df_onc.groupby(["hisse","kurum"])["oran2"].last().reset_index()
            onc.columns = ["hisse","kurum","onceki_oran"]
            tks2   = t2_sec.groupby("hisse")["tks2"].last().reset_index()
            artis_df = artis_df.merge(onc, on=["hisse","kurum"], how="left")
            artis_df = artis_df.merge(tks2, on="hisse", how="left")
            artis_df["onceki_oran"] = artis_df["onceki_oran"].fillna(0)
            artis_df["tks2"]        = artis_df["tks2"].fillna(0)
            artis_df["esik"] = artis_df["tks2"].apply(
                lambda x: 0.3 if x>=500_000_000 else (1.0 if x>=100_000_000 else 2.0)
            )
            artis_df = artis_df[
                (artis_df["onceki_oran"] < 0.5) &
                (artis_df["dolasim_pct"] >= artis_df["esik"]) &
                (artis_df["oran2"] <= 3)
            ].copy()
            baslik = "### 🔵 İlk Giriş Yapan Kurumlar (%0 → %3)"
        else:
            artis_df = artis_df[artis_df["dolasim_pct"] >= max(3, min_pct)].copy()
            baslik   = "### 📈 Seçilen Dönemde %3+ Artış Yapan Kurumlar"

        if not artis_df.empty:
            st.markdown(baslik)
            kl   = sorted(artis_df["kurum"].unique())
            cols = st.columns(min(len(kl), 4))
            for i, kurum in enumerate(kl):
                kdf  = artis_df[artis_df["kurum"]==kurum].sort_values(
                    "dolasim_pct", ascending=False).head(5)
                renk = "#1A5276" if kurum in AKILLI_PARA else "#1A7A3E"
                with cols[i % 4]:
                    st.markdown(
                        f"<div style='font-weight:bold;color:{renk};margin-bottom:4px;'>"
                        f"📈 {kurum}</div>", unsafe_allow_html=True
                    )
                    for _, r in kdf.iterrows():
                        st.markdown(
                            f"<div style='font-size:12px;padding:1px 0;'>"
                            f"<b>{r['hisse']}</b> "
                            f"<span style='color:{renk};font-weight:bold;'>"
                            f"+{r['dolasim_pct']:.1f}%</span></div>",
                            unsafe_allow_html=True
                        )

    st.markdown("---")

    # ── Toplama Alarm Listesi (alan + satan + sinyal) ─────────────────────────
    st.markdown("### 🚨 Toplama Alarm Listesi")

    # Tüm hisse+kurum kümülatif toplamları hesapla
    from takas_depo import _oku as _t2_alarm
    df_alarm_raw = _t2_alarm()
    if not df_alarm_raw.empty:
        df_alarm_raw = df_alarm_raw[df_alarm_raw["donem"].isin(secili_donemler)].copy()
        if "short_kapama" in df_alarm_raw.columns:
            mask = df_alarm_raw["short_kapama"] == True
            df_alarm_raw.loc[mask, "dolasim_pct"] = df_alarm_raw.loc[mask, "oran2"]

    if df_alarm_raw.empty:
        st.caption("Bu dönemde alarm yok.")
    else:
        # Her hisse+kurum kümülatif topla
        net = df_alarm_raw.groupby(["hisse", "kurum"]).agg(
            dolasim_pct=("dolasim_pct", "sum"),
            oran2=("oran2", "last"),
            tks2=("tks2", "last"),
        ).reset_index()
        net["dolasim_pct"] = net["dolasim_pct"].round(2)

        # Hisse bazında alanlar ve satanlar
        hisse_ozet = {}
        for hisse, grp_df in net.groupby("hisse"):
            alanlar = grp_df[grp_df["dolasim_pct"] >= min_pct].sort_values("dolasim_pct", ascending=False)
            satanlar = grp_df[grp_df["dolasim_pct"] <= -min_pct].sort_values("dolasim_pct")

            if alanlar.empty:
                continue

            toplam_alan  = alanlar["dolasim_pct"].sum()
            toplam_satan = abs(satanlar["dolasim_pct"].sum())
            alan_kurum   = len(alanlar)

            # Alarm seviyesi
            if toplam_alan >= 10:
                alarm = "🔴 KRİTİK"
                renk  = "#C0392B"
            elif toplam_alan >= 5:
                alarm = "🟠 GÜÇLÜ"
                renk  = "#E67E22"
            else:
                continue  # min_pct altı gösterme

            # ── Tks(2) değişimi ───────────────────────────────────────────────
            h_df_donem = df_alarm_raw[
                (df_alarm_raw["hisse"] == hisse) &
                (df_alarm_raw["tip"] == "gunluk")
            ].copy()
            donem_sira = sorted(h_df_donem["donem"].unique())
            tks2_bas = h_df_donem[h_df_donem["donem"] == donem_sira[0]]["tks2"].mean() if len(donem_sira) >= 1 else 0
            tks2_son = h_df_donem[h_df_donem["donem"] == donem_sira[-1]]["tks2"].mean() if len(donem_sira) >= 1 else 0
            tks2_degisim_pct = ((tks2_son - tks2_bas) / tks2_bas * 100) if tks2_bas > 0 else 0










            # ── Sinyal tespiti ────────────────────────────────────────────────
            satan_kurum = len(satanlar)
            satan_var   = not satanlar.empty and toplam_satan >= min_pct

            if tks2_degisim_pct > 0:
                # FD arttı → yeni hisse çıkmış
                sinyal = f"⚠️ Takas FD Değişti (+{tks2_degisim_pct:.0f}%)"
                sinyal_renk = "#7D6608"
            elif satan_var:
                # FD sabit, bizim kurumlardan satan var → Mal Devri
                if alan_kurum <= 2 and satan_kurum >= 5:
                    # Az kurum alıyor, çok kurum satıyor → Birikim
                    sinyal = "🎯 Birikim / Toplama"
                    sinyal_renk = "#1A5276"
                elif alan_kurum >= 5 and satan_kurum <= 2:
                    # Çok kurum alıyor, az kurum satıyor → Toplu Dağıtım
                    sinyal = "📤 Toplu Dağıtım"
                    sinyal_renk = "#C0392B"
                else:
                    sinyal = "🔄 Mal Devri"
                    sinyal_renk = "#1A5276"
            else:
                # FD sabit, bizim kurumlardan satan yok → piyasadan alıyor
                sinyal = "📈 Birikim"
                sinyal_renk = "#1A7A3E"

            hisse_ozet[hisse] = {
                "alarm": alarm, "renk": renk,
                "sinyal": sinyal, "sinyal_renk": sinyal_renk,
                "alanlar": alanlar, "satanlar": satanlar,
                "toplam_alan": toplam_alan, "toplam_satan": toplam_satan,
            }

        if not hisse_ozet:
            st.caption(f"Min %{min_pct} eşiğini geçen alarm yok.")
        else:
            # Kritik önce sırala
            sira_map = {"🔴 KRİTİK": 0, "🟠 GÜÇLÜ": 1}
            items = sorted(hisse_ozet.items(),
                           key=lambda x: (sira_map.get(x[1]["alarm"], 9), -x[1]["toplam_alan"]))

            cols3 = st.columns(3)
            for i, (hisse, data) in enumerate(items[:36]):
                renk  = data["renk"]
                sinyal_renk = data["sinyal_renk"]

                # Alan satırları
                alan_html = "".join([
                    f"<div style='font-size:11px;padding:1px 0 1px 6px;'>"
                    f"📥 <b style='color:{renk};'>{r['kurum']}</b> "
                    f"T2:%{r['oran2']:.1f} · <b style='color:#1A7A3E;'>+{r['dolasim_pct']:.2f}%</b></div>"
                    for _, r in data["alanlar"].iterrows()
                ])

                # Satan satırları
                if not data["satanlar"].empty:
                    satan_html = "".join([
                        f"<div style='font-size:11px;padding:1px 0 1px 6px;'>"
                        f"📤 <b style='color:#888;'>{r['kurum']}</b> "
                        f"T2:%{r['oran2']:.1f} · <b style='color:#C0392B;'>{r['dolasim_pct']:.2f}%</b></div>"
                        for _, r in data["satanlar"].iterrows()
                    ])
                else:
                    satan_html = "<div style='font-size:10px;color:#aaa;padding-left:6px;'>📤 Satan yok</div>"

                with cols3[i % 3]:
                    baslik = f"{hisse}  {data['alarm']}  · {data['sinyal']}"
                    with st.expander(baslik, expanded=False):
                        st.markdown(alan_html, unsafe_allow_html=True)
                        st.markdown("<hr style='margin:4px 0;border-color:#ddd;'>", unsafe_allow_html=True)
                        st.markdown(satan_html, unsafe_allow_html=True)
                        net_str = f"Net Alan: **+{data['toplam_alan']:.1f}%**"
                        if data['toplam_satan'] > 0:
                            net_str += f"  |  Net Satan: **-{data['toplam_satan']:.1f}%**"
                        st.caption(net_str)

    st.markdown("---")

    # ── Kurum Bazlı Detay Tablo ───────────────────────────────────────────────
    st.markdown("### 📋 Kurum Bazlı Detay Tablo")
    pivot_df = takas_analiz(secili_donemler, mod)

    if not pivot_df.empty:
        k_cols = [c for c in pivot_df.columns if c not in ["hisse","tks2"]]
        pivot_df["TOPLAM"] = pivot_df[k_cols].sum(axis=1).round(2)
        pivot_df = pivot_df[pivot_df["TOPLAM"].abs() >= min_pct]
        pivot_df = pivot_df.sort_values("TOPLAM", ascending=False)

        if pivot_df.empty:
            st.caption(f"Min %{min_pct} geçen hisse yok.")
        else:
            fmt = {c:"{:+.2f}" for c in k_cols + ["TOPLAM"]}
            fmt["tks2"] = "{:,.0f}"

            def _r(v):
                if not isinstance(v, float): return "color:#888"
                if v >= 1.0:  return "background-color:#D5F5E3;color:#1A5276;font-weight:bold"
                if v > 0:     return "color:#1A5276;font-weight:bold"
                if v <= -1.0: return "background-color:#FADBD8;color:#C0392B;font-weight:bold"
                if v < 0:     return "color:#C0392B"
                return "color:#888"

            st.dataframe(
                pivot_df.style.map(_r, subset=k_cols+["TOPLAM"]).format(fmt, na_rep="—"),
                use_container_width=True, height=500
            )
            st.download_button(
                "⬇️ Excel İndir",
                data=_takas_excel_indir(pivot_df, secili_donemler),
                file_name=f"takas_analiz_{mod}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="takas_excel_btn"
            )



def takas_veri_yukle_bolumu():
    """Veri Yükle sekmesindeki takas bölümü."""
    st.markdown("### 🏦 Kurum Takas Dosyaları")
    st.caption(
        "Dosya adı formatı: **KURUM_AY_DONEM.xlsx**  |  "
        "Aylık: `TERA__2026_01_03.xlsx`  |  "
        "Haftalık: `TERA_202604_01.xlsx`  |  "
        "Günlük: `TERA_20260420.xlsx`"
    )

    with st.form("takas_yukle_form"):
        yuklenen = st.file_uploader(
            "📂 Dosyaları Seçin (birden fazla seçilebilir):",
            type=["xlsx"],
            accept_multiple_files=True,
            key="takas_dosya_yukle"
        )
        yukle_btn = st.form_submit_button(
            "✅ Yükle", use_container_width=True
        )

    if yukle_btn:
        if not yuklenen:
            st.error("En az 1 dosya seçin!")
        else:
            dosya_listesi = [(f.name, f) for f in yuklenen]
            eklenen, hatalar = dosyalar_yukle(dosya_listesi)

            if eklenen:
                for msg in eklenen:
                    st.success(msg)
            if hatalar:
                for msg in hatalar:
                    st.warning(msg)

            if eklenen:
                st.rerun()

    # Kayıtlı dönemler
    st.markdown("---")
    st.markdown("**📂 Kayıtlı Veriler:**")

    col_g, col_h, col_a = st.columns(3)

    for col, tip, label in [
        (col_g, "gunluk", "📅 Günlük"),
        (col_h, "haftalik", "📆 Haftalık"),
        (col_a, "aylik", "🗓️ Aylık")
    ]:
        with col:
            st.markdown(f"**{label}:**")
            donemler = donemler_listele(tip)
            if donemler:
                kurumlar = kurumlar_listele()
                for d in donemler[:10]:
                    c1, c2 = st.columns([3, 1])
                    c1.markdown(f"`{d}`")
                    if c2.button("🗑️", key=f"takas_sil_{tip}_{d}"):
                        for k in kurumlar:
                            donem_sil(k, d)
                        st.rerun()
            else:
                st.caption("Henüz yok")


def _takas_excel_indir(df: pd.DataFrame, donemler: list):
    """Takas tablosunu Excel'e dönüştürür."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Takas Analiz"

    cols = list(df.columns)
    tarih = datetime.now().strftime("%d.%m.%Y")

    # Başlık
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws["A1"]
    c.value = f"Takas Analizi | {', '.join(donemler)} | {tarih}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="1A252F")
    c.alignment = Alignment(horizontal="center")

    # Header
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=i, value=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="1A3A5C")
        cell.alignment = Alignment(horizontal="center")

    # Veri
    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 3
        for ci, col in enumerate(cols, 1):
            v = row[col]
            cell = ws.cell(row=r, column=ci, value=v)
            if isinstance(v, float):
                cell.number_format = "+0.00;-0.00;0.00"
                if v >= 1.0:
                    cell.font = Font(bold=True, color="1A5276")
                elif v > 0:
                    cell.font = Font(color="1A5276")
                elif v <= -1.0:
                    cell.font = Font(bold=True, color="C0392B")
                elif v < 0:
                    cell.font = Font(color="C0392B")

    # Kolon genişlikleri
    ws.column_dimensions["A"].width = 10
    for i in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.freeze_panes = "B3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
