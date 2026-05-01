"""
mkk_app.py — MKK Oran Analizi sekmesi
app.py'ye import edilir.

Dosya adı formatı:
  Haftalık: MKK_202604_01.xlsx  → dönem: 202604_01
  Aylık:    MKK_2026_04.xlsx    → dönem: 2026_04
"""

import streamlit as st
import pandas as pd
import re
from mkk_depo import yukle, sil, donemler_listele, pivot_olustur, veri_getir
from io import BytesIO


def _tip_ve_donem_bul(dosya_adi: str):
    """
    Dosya adından tip ve dönem çıkarır.
    MKK_202604_01.xlsx → ('haftalik', '202604_01')
    MKK_2026_04.xlsx   → ('aylik',    '2026_04')
    """
    stem = dosya_adi.upper().replace(".XLSX", "").replace(".xlsx", "")
    # Haftalık: MKK_YYYYMM_HH
    m = re.match(r".*MKK[_\-](\d{6})[_\-](\d{2})$", stem)
    if m:
        return "haftalik", f"{m.group(1)}_{m.group(2)}"
    # Aylık: MKK_YYYY_MM
    m = re.match(r".*MKK[_\-](\d{4})[_\-](\d{2})$", stem)
    if m:
        return "aylik", f"{m.group(1)}_{m.group(2)}"
    return None, None


def _renk(val):
    if isinstance(val, (int, float)):
        if val > 0: return "color:#1A5276;font-weight:bold"
        if val < 0: return "color:#C0392B;font-weight:bold"
    return "color:#888888"


def _tablo_goster(pivot: pd.DataFrame, donem_cols: list):
    fmt = {c: "{:+.2f}" for c in donem_cols + ["KÜMÜLATİF"]}
    styled = pivot.style.map(_renk, subset=donem_cols + ["KÜMÜLATİF"]).format(fmt, na_rep="—")
    st.dataframe(styled, hide_index=True, use_container_width=True, height=600)


def _excel_indir(pivot: pd.DataFrame, donem_cols: list, baslik: str) -> BytesIO:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MKK Analiz"
    cols = list(pivot.columns)

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws["A1"]
    c.value = baslik
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="1A252F")
    c.alignment = Alignment(horizontal="center")

    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=i, value=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="1A3A5C")
        cell.alignment = Alignment(horizontal="center")

    for ri, (_, row) in enumerate(pivot.iterrows()):
        r = ri + 3
        for ci, col in enumerate(cols, 1):
            v = row[col]
            cell = ws.cell(row=r, column=ci, value=v)
            if isinstance(v, float):
                cell.number_format = "+0.00;-0.00;0.00"
                if v > 0: cell.font = Font(bold=True, color="1A5276")
                elif v < 0: cell.font = Font(bold=True, color="C0392B")

    ws.column_dimensions["A"].width = 10
    for i in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.freeze_panes = "B3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _veri_yukle_bolumu():
    """Veri Yükle sekmesinden çağrılacak MKK yükleme bölümü."""
    st.markdown("### 📊 MKK Dosyaları")
    st.caption(
        "Dosya adı formatı: "
        "**Haftalık:** `MKK_202604_01.xlsx` | "
        "**Aylık:** `MKK_2026_04.xlsx`"
    )

    with st.form("mkk_toplu_yukle"):
        dosyalar = st.file_uploader(
            "📂 Dosyaları Seçin (birden fazla):",
            type=["xlsx"],
            accept_multiple_files=True,
            key="mkk_dosya_yukle"
        )
        btn = st.form_submit_button("✅ Yükle", use_container_width=True)

    if btn:
        if not dosyalar:
            st.error("En az 1 dosya seçin!")
        else:
            for dosya in dosyalar:
                tip, donem = _tip_ve_donem_bul(dosya.name)
                if tip is None:
                    st.warning(f"⚠️ `{dosya.name}` — format tanınamadı, atlandı.")
                    continue
                ok, msg = yukle(donem, tip, dosya)
                if ok:
                    st.success(f"`{dosya.name}` → {msg}")
                else:
                    st.warning(f"`{dosya.name}` → {msg}")
            st.rerun()

    # Kayıtlı dönemler
    st.markdown("---")
    col_h, col_a = st.columns(2)

    with col_h:
        st.markdown("**📅 Haftalık:**")
        haf_don = donemler_listele("haftalik")
        if haf_don:
            for d in sorted(haf_don, reverse=True):
                c1, c2 = st.columns([4, 1])
                c1.markdown(f"`{d}`")
                if c2.button("🗑️", key=f"mkk_sil_h_{d}"):
                    sil(d, "haftalik")
                    st.rerun()
        else:
            st.caption("Henüz yok")

    with col_a:
        st.markdown("**📆 Aylık:**")
        ayl_don = donemler_listele("aylik")
        if ayl_don:
            for d in sorted(ayl_don, reverse=True):
                c1, c2 = st.columns([4, 1])
                c1.markdown(f"`{d}`")
                if c2.button("🗑️", key=f"mkk_sil_a_{d}"):
                    sil(d, "aylik")
                    st.rerun()
        else:
            st.caption("Henüz yok")


def _analiz_panel(tip: str, baslik: str, kolon: str):
    donemler = donemler_listele(tip)

    if not donemler:
        st.info(f"📂 {baslik} verisi yok. **Veri Yükle** sekmesinden yükleyin.")
        return

    c1, c2, c3 = st.columns([4, 1, 1])
    with c1:
        secili = st.multiselect(
            "Dönemler:", sorted(donemler, reverse=True),
            default=sorted(donemler, reverse=True)[:4],
            key=f"mkk_sec_{tip}"
        )
    with c2:
        min_pct = st.number_input(
            "Min Kümülatif:", value=0.0, step=0.5,
            key=f"mkk_min_{tip}"
        )
    with c3:
        filtre = st.selectbox(
            "Filtre:", ["Tümü", "🚀 Sürekli Artan", "🟢 Pozitif", "🔴 Negatif"],
            key=f"mkk_filtre_{tip}"
        )

    if not secili:
        st.warning("Dönem seçin.")
        return

    pivot = pivot_olustur(secili, tip, kolon)
    if pivot.empty:
        st.warning("Veri bulunamadı.")
        return

    donem_cols = [c for c in pivot.columns if c not in ["hisse", "KÜMÜLATİF", "TREND"]]

    if min_pct != 0.0:
        pivot = pivot[pivot["KÜMÜLATİF"].abs() >= min_pct]
    if filtre == "🚀 Sürekli Artan":
        pivot = pivot[pivot["TREND"] == "🚀"]
    elif filtre == "🟢 Pozitif":
        pivot = pivot[pivot["KÜMÜLATİF"] > 0]
    elif filtre == "🔴 Negatif":
        pivot = pivot[pivot["KÜMÜLATİF"] < 0]

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Toplam Hisse", len(pivot))
    m2.metric("🚀 Sürekli Artan", len(pivot[pivot["TREND"] == "🚀"]))
    m3.metric("🟢 Pozitif", len(pivot[pivot["KÜMÜLATİF"] > 0]))
    m4.metric("🔴 Negatif", len(pivot[pivot["KÜMÜLATİF"] < 0]))

    st.divider()
    _tablo_goster(pivot, donem_cols)

    buf = _excel_indir(pivot, donem_cols, f"{baslik} | {', '.join(secili)}")
    st.download_button(
        "⬇️ Excel İndir", data=buf,
        file_name=f"mkk_{tip}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def mkk_sekmesi():
    alt_tab1, alt_tab2, alt_tab3 = st.tabs([
        "📅 Haftalık",
        "📆 Aylık",
        "📊 Kurum Oran Tablosu",
    ])

    with alt_tab1:
        st.markdown("### 📅 Haftalık MKK — Son Tarih (I) − İlk Tarih (E)")
        _analiz_panel("haftalik", "Haftalık MKK", "pp_fark")

    with alt_tab2:
        st.markdown("### 📆 Aylık MKK — Son Tarih (I) − İlk Tarih (E)")
        _analiz_panel("aylik", "Aylık MKK", "pp_fark")

    with alt_tab3:
        st.markdown("### 📊 Kurum Oran Tablosu — Fark Kurum Oran (M Kolonu)")
        st.caption("Haftalık ve Aylık dönemler birlikte seçilebilir.")

        haf_don = donemler_listele("haftalik")
        ayl_don = donemler_listele("aylik")
        tum_don = sorted(
            [f"H:{d}" for d in haf_don] + [f"A:{d}" for d in ayl_don],
            reverse=True
        )

        if not tum_don:
            st.info("📂 Veri yok. Veri Yükle sekmesinden yükleyin.")
            return

        c1, c2 = st.columns([4, 2])
        with c1:
            secili = st.multiselect(
                "Dönemler (H=Haftalık, A=Aylık):", tum_don,
                default=tum_don[:4], key="mkk_kur_sec"
            )
        with c2:
            filtre_k = st.selectbox(
                "Filtre:", ["Tümü", "🚀 Sürekli Artan", "🟢 Pozitif", "🔴 Negatif"],
                key="mkk_kur_filtre"
            )

        if not secili:
            st.warning("Dönem seçin.")
            return

        haf_sec = [d[2:] for d in secili if d.startswith("H:")]
        ayl_sec = [d[2:] for d in secili if d.startswith("A:")]

        parcalar = []
        if haf_sec:
            df_h = veri_getir(haf_sec, "haftalik")
            if not df_h.empty:
                parcalar.append(df_h)
        if ayl_sec:
            df_a = veri_getir(ayl_sec, "aylik")
            if not df_a.empty:
                parcalar.append(df_a)

        if not parcalar:
            st.warning("Veri bulunamadı.")
            return

        df_tum = pd.concat(parcalar, ignore_index=True)
        pivot = df_tum.pivot_table(
            index="hisse", columns="donem", values="kur_oran_fark", aggfunc="last"
        ).reset_index()
        pivot.columns.name = None
        donem_cols = sorted([c for c in pivot.columns if c != "hisse"])
        pivot["KÜMÜLATİF"] = pivot[donem_cols].sum(axis=1).round(2)

        def trend(row):
            vals = [row[c] for c in donem_cols if pd.notna(row[c])]
            if len(vals) < 2: return "—"
            son3 = vals[-3:] if len(vals) >= 3 else vals
            if all(v > 0 for v in son3) and all(son3[i] > son3[i-1] for i in range(1, len(son3))):
                return "🚀"
            if all(v > 0 for v in vals): return "🟢"
            if sum(1 for v in vals if v > 0) > sum(1 for v in vals if v < 0): return "🟡"
            return "🔴"

        pivot["TREND"] = pivot.apply(trend, axis=1)
        pivot = pivot.sort_values("KÜMÜLATİF", ascending=False).reset_index(drop=True)

        if filtre_k == "🚀 Sürekli Artan":
            pivot = pivot[pivot["TREND"] == "🚀"]
        elif filtre_k == "🟢 Pozitif":
            pivot = pivot[pivot["KÜMÜLATİF"] > 0]
        elif filtre_k == "🔴 Negatif":
            pivot = pivot[pivot["KÜMÜLATİF"] < 0]

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Toplam Hisse", len(pivot))
        m2.metric("🚀 Sürekli Artan", len(pivot[pivot["TREND"] == "🚀"]))
        m3.metric("🟢 Pozitif", len(pivot[pivot["KÜMÜLATİF"] > 0]))
        m4.metric("🔴 Negatif", len(pivot[pivot["KÜMÜLATİF"] < 0]))

        st.divider()
        _tablo_goster(pivot, donem_cols)

        buf = _excel_indir(pivot, donem_cols, f"Kurum Oran | {', '.join(secili)}")
        st.download_button(
            "⬇️ Excel İndir", data=buf,
            file_name="mkk_kurum_oran.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
