"""
endeks_takip.py — BIST 30 Endeks Sıralama Takip Modülü
Haftalık beta_endeks.xlsx dosyasından sıralama değişimini takip eder.
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as plt_go
from pathlib import Path
import openpyxl

BETA_FILE    = Path("data/BETA_ENDEKS.xlsx")
BIST_FD      = Path("data/bist_fd.xlsx")
BIST30_FILE  = Path("data/Bist_30.xlsx")
YILDIZ_FILE  = Path("data/Bist_yildiz_pazar.xlsx")


def _yildiz_pazar_hisseler() -> set:
    """Yıldız Pazar hisse listesini okur."""
    if not YILDIZ_FILE.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(YILDIZ_FILE, read_only=True)
        ws = wb.active
        return {str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
    except:
        return set()

# BIST 30 üyeleri — dosyadan oku, yoksa sabit liste
def _bist30_uyeleri() -> set:
    if BIST30_FILE.exists():
        try:
            wb = openpyxl.load_workbook(BIST30_FILE, read_only=True)
            ws = wb.active
            return {str(row[0]).strip() for row in ws.iter_rows(min_row=2, values_only=True) if row[0]}
        except:
            pass
    # Fallback sabit liste
    return {
        "AKBNK","ASELS","ASTOR","BIMAS","DSTKF","EKGYO","ENKAI","EREGL",
        "FROTO","GARAN","GUBRF","ISCTR","KCHOL","KRDMD","MGROS","PETKM",
        "PGSUS","SAHOL","SASA","SISE","TAVHL","TCELL","THYAO","TOASO",
        "TRALT","TTKOM","TUPRS","VAKBN","YKBNK","AEFES"
    }

def _bist_hisseler() -> set:
    """bist_fd.xlsx'ten geçerli hisse listesini okur."""
    if not BIST_FD.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(BIST_FD, read_only=True)
        ws = wb.active
        hisseler = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                hisseler.add(str(row[0]).strip())
        return hisseler
    except:
        return set()


def _oku_beta(yildiz_only: bool = True) -> dict:
    """BETA_ENDEKS.xlsx dosyasını okur. {sheet_name: DataFrame}"""
    if not BETA_FILE.exists():
        return {}

    bist_hisseler   = _bist_hisseler()
    yildiz_hisseler = _yildiz_pazar_hisseler() if yildiz_only else set()
    wb = openpyxl.load_workbook(BETA_FILE, read_only=True)
    sonuc = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        kayitlar = []
        for row in rows[1:]:
            if row[1] is None:
                continue
            senet = str(row[1]).strip()
            if not senet:
                continue
            # BIST FD filtresi
            if bist_hisseler and senet not in bist_hisseler:
                continue
            # Yıldız Pazar filtresi
            if yildiz_hisseler and senet not in yildiz_hisseler:
                continue
            try:
                etki = float(str(row[2]).replace(',', '.')) if row[2] else 0.0
                beta = float(str(row[3]).replace(',', '.')) if row[3] else 0.0
                pay_adet = int(row[4]) if row[4] else 0
                fiyat_str = str(row[5]).replace('.', '').replace(',', '.') if row[5] else '0'
                fiyat = float(fiyat_str) if fiyat_str else 0.0
            except:
                continue

            kayitlar.append({
                'senet': senet,
                'endeks_etkisi': etki,
                'beta': beta,
                'pay_adet': pay_adet,
                'fiyat': fiyat,
            })

        if kayitlar:
            df = pd.DataFrame(kayitlar)
            df = df.sort_values('endeks_etkisi', ascending=False).reset_index(drop=True)
            df['sira'] = df.index + 1
            sonuc[sheet_name] = df

    return sonuc


def _sira_degisim(veriler: dict) -> pd.DataFrame:
    """Tüm dönemlerdeki sıralamayı birleştirir."""
    donemler = list(veriler.keys())
    if not donemler:
        return pd.DataFrame()

    tum_hisseler = set()
    for df in veriler.values():
        tum_hisseler.update(df['senet'].tolist())

    kayitlar = []
    for hisse in sorted(tum_hisseler):
        satir = {'senet': hisse}
        for donem in donemler:
            df = veriler[donem]
            h = df[df['senet'] == hisse]
            satir[donem] = int(h['sira'].values[0]) if not h.empty else None
        kayitlar.append(satir)

    return pd.DataFrame(kayitlar)


def endeks_takip_sekme():
    """Streamlit sekme içeriği."""
    st.subheader("📊 BIST 30 Endeks Sıralama Takibi")

    BIST30_UYELERI = _bist30_uyeleri()

    # Yıldız Pazar filtresi seçimi
    col_f1, col_f2 = st.columns([2, 4])
    with col_f1:
        yildiz_filtre = st.radio(
            "Hisse Evreni:",
            ["🌟 Yıldız Pazar", "📋 BIST Tüm"],
            horizontal=True,
            key="endeks_yildiz_sec"
        )

    # Dosya yükleme
    if not BETA_FILE.exists():
        st.info("Henüz veri yüklenmemiş.")
        yukle = st.file_uploader("BETA_ENDEKS.xlsx yükle", type=["xlsx"], key="beta_yukle")
        if yukle:
            BETA_FILE.parent.mkdir(parents=True, exist_ok=True)
            with open(BETA_FILE, "wb") as f:
                f.write(yukle.read())
            st.success("✅ Dosya yüklendi!")
            st.rerun()
        return

    veriler = _oku_beta(yildiz_only=(yildiz_filtre == "🌟 Yıldız Pazar"))
    if not veriler:
        st.warning("Dosya okunamadı.")
        return

    donemler = list(veriler.keys())
    son_donem = donemler[-1]
    son_df = veriler[son_donem]

    # ── Özet metrikler ──────────────────────────────────────────────────────
    evre_label = "🌟 Yıldız Pazar" if yildiz_filtre == "🌟 Yıldız Pazar" else "📋 BIST Tüm"
    st.caption(f"Son dönem: **{son_donem}** | {len(donemler)} hafta | {len(son_df)} hisse | {evre_label}")

    # ── Son dönem sıralaması + önceki dönemle karşılaştırma ─────────────────
    if len(donemler) >= 2:
        onceki_donem = donemler[-2]
        onceki_df = veriler[onceki_donem]

        son_df = son_df.merge(
            onceki_df[['senet', 'sira']].rename(columns={'sira': 'onceki_sira'}),
            on='senet', how='left'
        )
        son_df['sira_degisim'] = son_df['onceki_sira'] - son_df['sira']  # pozitif = yükseldi
    else:
        son_df['onceki_sira'] = None
        son_df['sira_degisim'] = None

    # ── Alarm bölgeleri ──────────────────────────────────────────────────────
    col_gir, col_kri, col_cik = st.columns(3)

    giris_adayi = son_df[(son_df['sira'] <= 30) & 
                          (~son_df['senet'].isin(BIST30_UYELERI))]
    
    kritis_bolge = son_df[(son_df['sira'] >= 25) & (son_df['sira'] <= 35)]
    
    cikis_adayi = son_df[(son_df['sira'] > 30) & 
                          (son_df['senet'].isin(BIST30_UYELERI))]

    with col_gir:
        st.markdown("### 🚀 Giriş Adayları")
        if not giris_adayi.empty:
            for _, r in giris_adayi.iterrows():
                uye = r['senet'] in BIST30_UYELERI
                renk = "#888888" if uye else "#1A7A3E"
                etiket = "✅ Zaten Üye" if uye else "🆕 Yeni Giriş"
                st.markdown(
                    f"<div style='border-left:4px solid {renk};padding:6px 10px;"
                    f"margin:4px 0;background:#F0FFF4;border-radius:0 4px 4px 0;'>"
                    f"<b>{r['senet']}</b> → #{int(r['sira'])}"
                    f"<span style='color:{renk};margin-left:8px;'>↑ {int(r['onceki_sira']) if pd.notna(r.get('onceki_sira')) else '?'}→{int(r['sira'])}</span>"
                    f"<span style='margin-left:8px;font-size:11px;'>{etiket}</span>"
                    f"<br><span style='font-size:11px;color:#555;'>Etki: %{r['endeks_etkisi']:.2f} | Beta: {r['beta']:.2f}</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
        else:
            st.caption("Bu hafta giriş yok")

    with col_kri:
        st.markdown("### ⚠️ Kritik Bölge (25-35)")
        kritis_goster = kritis_bolge.head(12)
        for _, r in kritis_goster.iterrows():
            degisim = ""
            renk = "#E67E22"
            if pd.notna(r.get('sira_degisim')):
                d = int(r['sira_degisim'])
                if d > 0:
                    degisim = f" ↑{d}"
                    renk = "#1A7A3E"
                elif d < 0:
                    degisim = f" ↓{abs(d)}"
                    renk = "#C0392B"
            uye_badge = " 🏅" if r['senet'] in BIST30_UYELERI else ""
            st.markdown(
                f"<div style='padding:3px 8px;margin:2px 0;font-size:13px;'>"
                f"<b>#{int(r['sira'])}</b> {r['senet']}{uye_badge}"
                f"<span style='color:{renk};margin-left:6px;'>{degisim}</span>"
                f"</div>",
                unsafe_allow_html=True
            )

    with col_cik:
        st.markdown("### ⬇️ Çıkış Adayları")
        if not cikis_adayi.empty:
            for _, r in cikis_adayi.iterrows():
                uye = r['senet'] in BIST30_UYELERI
                renk = "#C0392B"
                etiket = "⚠️ Üye Çıkıyor!" if uye else "📉 Çıkış"
                st.markdown(
                    f"<div style='border-left:4px solid {renk};padding:6px 10px;"
                    f"margin:4px 0;background:#FFF5F5;border-radius:0 4px 4px 0;'>"
                    f"<b>{r['senet']}</b> → #{int(r['sira'])}"
                    f"<span style='color:{renk};margin-left:8px;'>↓ {int(r['onceki_sira']) if pd.notna(r['onceki_sira']) else '?'}→{int(r['sira'])}</span>"
                    f"<span style='margin-left:8px;font-size:11px;'>{etiket}</span>"
                    f"<br><span style='font-size:11px;color:#555;'>Etki: %{r['endeks_etkisi']:.2f} | Beta: {r['beta']:.2f}</span>"
                    f"</div>",
                    unsafe_allow_html=True
                )
        else:
            st.caption("Bu hafta çıkış yok")

    st.divider()

    # ── Tarihsel sıralama grafiği ────────────────────────────────────────────
    st.markdown("### 📈 Hisse Sıralama Değişimi")
    
    sira_df = _sira_degisim(veriler)
    
    col_s1, col_s2 = st.columns([1, 3])
    with col_s1:
        secili_hisse = st.selectbox(
            "Hisse seç:",
            sorted(sira_df['senet'].tolist()),
            key="endeks_hisse_sec"
        )

    if secili_hisse:
        h_df = sira_df[sira_df['senet'] == secili_hisse].iloc[0]
        siralar = [(d, h_df[d]) for d in donemler if pd.notna(h_df.get(d))]

        if siralar:
            x_vals = [s[0] for s in siralar]
            y_vals = [s[1] for s in siralar]

            fig = plt_go.Figure()
            fig.add_trace(plt_go.Scatter(
                x=x_vals, y=y_vals,
                mode='lines+markers+text',
                text=[str(int(v)) for v in y_vals],
                textposition='top center',
                line=dict(color='#1A5276', width=2),
                marker=dict(size=8)
            ))
            # 30. sıra çizgisi
            fig.add_hline(y=30, line_dash="dash", line_color="#C0392B",
                         annotation_text="30. Sıra (Endeks Sınırı)")
            fig.add_hline(y=25, line_dash="dot", line_color="#E67E22",
                         annotation_text="25. Sıra (Giriş Eşiği)")
            fig.add_hline(y=35, line_dash="dot", line_color="#E67E22",
                         annotation_text="35. Sıra (Çıkış Eşiği)")

            fig.update_layout(
                title=f"{secili_hisse} — Endeks Sıralama Değişimi",
                yaxis=dict(autorange='reversed', title='Sıra'),
                height=400,
                margin=dict(l=10, r=10, t=40, b=10),
                plot_bgcolor='#FAFAFA', paper_bgcolor='white'
            )
            st.plotly_chart(fig, use_container_width=True)

    st.divider()

    # ── Son dönem tam liste ───────────────────────────────────────────────────
    st.markdown(f"### 📋 Son Dönem Tam Liste ({son_donem})")
    
    goster_df = son_df[['sira', 'senet', 'endeks_etkisi', 'beta', 
                          'pay_adet', 'fiyat', 'onceki_sira', 'sira_degisim']].copy()
    goster_df.columns = ['Sıra', 'Senet', 'Endeks Etkisi %', 'Beta', 
                          'Pay Adedi', 'Fiyat', 'Önceki Sıra', 'Değişim']

    def renk_sira(val):
        if isinstance(val, (int, float)):
            if val <= 25: return 'background-color:#D5F5E3;font-weight:bold'
            if val <= 30: return 'color:#1A7A3E;font-weight:bold'
            if val <= 35: return 'color:#E67E22'
            if val > 35: return 'color:#C0392B'
        return ''

    def renk_degisim(val):
        if isinstance(val, (int, float)):
            if val > 0: return 'color:#1A7A3E;font-weight:bold'
            if val < 0: return 'color:#C0392B;font-weight:bold'
        return ''

    styled = goster_df.style \
        .map(renk_sira, subset=['Sıra']) \
        .map(renk_degisim, subset=['Değişim']) \
        .format({
            'Endeks Etkisi %': '{:.2f}',
            'Beta': '{:.2f}',
            'Pay Adedi': '{:,.0f}',
            'Fiyat': '{:.2f}',
            'Değişim': lambda v: f'{int(v):+d}' if pd.notna(v) else '—'
        }, na_rep='—')

    st.dataframe(styled, hide_index=True, use_container_width=True, height=600)

    # ── Dosya güncelleme ──────────────────────────────────────────────────────
    with st.expander("📤 Dosyayı Güncelle"):
        yukle = st.file_uploader("Yeni BETA_ENDEKS.xlsx yükle", 
                                  type=["xlsx"], key="beta_guncelle")
        if yukle:
            with open(BETA_FILE, "wb") as f:
                f.write(yukle.read())
            st.success("✅ Güncellendi!")
            st.rerun()
