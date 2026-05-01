"""
app_takas.py — TAKAS ANALİZİ Sekmesi
app.py'ye import edilir.
"""

import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from takas_depo import (
    dosyalar_yukle, donemler_listele, kurumlar_listele,
    kurum_net_pozisyon, alarm_listesi, takas_analiz,
    hisse_kurum_detay, trend_analiz, donem_sil,
    BUYUK_YERLI, AKILLI_PARA, FON_YABANCI, KURUMLAR
)


def aylik_trend_analiz(df_csv: pd.DataFrame) -> dict:
    """6 aylık veriden sinyal tespiti."""
    if df_csv.empty:
        return {}
    df = df_csv[df_csv["tip"] == "aylik"].copy()
    if df.empty:
        return {}
    donemler = sorted(df["donem"].unique())
    if len(donemler) < 2:
        return {}

    sonuclar = {
        "sifirdan_giris": [], "buyuk_alim": [], "duzenli_artis": [],
        "sifira_cikis": [], "buyuk_satis": [], "duzenli_azalis": [],
    }

    for hisse in df["hisse"].unique():
        for kurum in df["kurum"].unique():
            k_df = df[(df["hisse"]==hisse) & (df["kurum"]==kurum)].sort_values("donem")
            if k_df.empty or len(k_df) < 2: continue
            oranlar = k_df["oran2"].tolist()
            d_list  = k_df["donem"].tolist()
            ilk = oranlar[0]; son = oranlar[-1]; fark = son - ilk
            kron = " → ".join([f"%{o:.1f}" for o in oranlar])

            if ilk < 0.1 and son >= 1.0:
                sonuclar["sifirdan_giris"].append({"Hisse":hisse,"Kurum":kurum,
                    "İlk":d_list[0],"Son":d_list[-1],"Son %":round(son,2),"Kronoloji":kron})

            for i in range(1, len(oranlar)):
                d_fark = oranlar[i] - oranlar[i-1]
                if d_fark >= 3.0:
                    sonuclar["buyuk_alim"].append({"Hisse":hisse,"Kurum":kurum,
                        "Dönem":d_list[i],"Önceki %":round(oranlar[i-1],2),
                        "Sonraki %":round(oranlar[i],2),"Artış":round(d_fark,2),"Kronoloji":kron})
                if d_fark <= -3.0:
                    sonuclar["buyuk_satis"].append({"Hisse":hisse,"Kurum":kurum,
                        "Dönem":d_list[i],"Önceki %":round(oranlar[i-1],2),
                        "Sonraki %":round(oranlar[i],2),"Düşüş":round(d_fark,2),"Kronoloji":kron})

            if len(oranlar) >= 3:
                son3 = oranlar[-3:]
                if all(son3[i]>son3[i-1] for i in range(1,len(son3))) and son3[-1]>=0.5:
                    sonuclar["duzenli_artis"].append({"Hisse":hisse,"Kurum":kurum,
                        "Son %":round(son,2),"Toplam":round(fark,2),"Kronoloji":kron})
                if all(son3[i]<son3[i-1] for i in range(1,len(son3))) and son3[0]>=0.5:
                    sonuclar["duzenli_azalis"].append({"Hisse":hisse,"Kurum":kurum,
                        "Son %":round(son,2),"Toplam":round(fark,2),"Kronoloji":kron})

            if ilk >= 1.0 and son < 0.1:
                sonuclar["sifira_cikis"].append({"Hisse":hisse,"Kurum":kurum,
                    "İlk %":round(ilk,2),"Kronoloji":kron})

    return {k: pd.DataFrame(v) for k, v in sonuclar.items()}


def _kart(hisse, kurum, renk, detay, kron):
    st.markdown(
        f"""<div style='border-left:4px solid {renk};padding:5px 10px;
        margin:3px 0;background:#FAFAFA;border-radius:0 4px 4px 0;'>
        <b>{hisse}</b> — <span style='color:{renk}'>{kurum}</span>
        <span style='float:right;font-weight:bold;color:{renk}'>{detay}</span><br>
        <small style='color:#666'>{kron}</small>
        </div>""", unsafe_allow_html=True
    )


def aylik_trend_goster():
    """Aylık trend — hisse bazlı, takip kurumları, %5+ filtresi."""
    from takas_depo import _oku, AKILLI_PARA, KURUMLAR

    df = _oku()
    if df.empty:
        st.info("Aylık veri yok.")
        return

    df_aylik = df[df["tip"] == "aylik"].copy()
    if df_aylik.empty:
        st.info("Aylık veri bulunamadı.")
        return

    donemler = sorted(df_aylik["donem"].unique())
    son_donem = donemler[-1]
    st.caption(f"📅 Mevcut aylar: {' | '.join(donemler)} | **Son: {son_donem}**")

    # ── Filtreler ─────────────────────────────────────────────────────────────
    col1, col2, col3 = st.columns([2, 1, 1])
    with col1:
        sec_ay = st.selectbox("Ay seç", donemler, index=len(donemler)-1, key="trend_ay")
    with col2:
        min_pct = st.number_input("Min pozisyon %", value=5.0, step=1.0, key="trend_min_pct")
    with col3:
        min_kurum = st.number_input("Min kurum sayısı", value=1, step=1, min_value=1, key="trend_min_kurum")

    # ── Seçili ay verisi ──────────────────────────────────────────────────────
    df_sec = df_aylik[df_aylik["donem"] == sec_ay].copy()

    # Sadece takip kurumları
    df_sec = df_sec[df_sec["kurum"].isin(KURUMLAR)]

    # Min %5 filtresi
    df_sec = df_sec[df_sec["oran2"] >= min_pct]

    if df_sec.empty:
        st.warning(f"'{sec_ay}' döneminde %{min_pct}+ pozisyonlu takip kurumu bulunamadı.")
        return

    # ── Önceki ay için trend hesapla ──────────────────────────────────────────
    onceki_ay = donemler[donemler.index(sec_ay)-1] if donemler.index(sec_ay) > 0 else None

    if onceki_ay:
        df_onc = df_aylik[df_aylik["donem"] == onceki_ay][["hisse","kurum","oran2"]].rename(
            columns={"oran2": "onceki_oran"}
        )
        df_sec = df_sec.merge(df_onc, on=["hisse","kurum"], how="left")
        df_sec["degisim"] = (df_sec["oran2"] - df_sec["onceki_oran"].fillna(0)).round(2)
        df_sec["yeni_giris"] = df_sec["onceki_oran"].isna() | (df_sec["onceki_oran"] < 1.0)
    else:
        df_sec["onceki_oran"] = 0
        df_sec["degisim"] = df_sec["oran2"]
        df_sec["yeni_giris"] = True

    # ── Hisse bazlı grupla ────────────────────────────────────────────────────
    rows = []
    for hisse, grp in df_sec.groupby("hisse"):
        grp = grp.sort_values("oran2", ascending=False)

        kurumlar_str = "  ".join([
            f"**{r['kurum']}** %{r['oran2']:.1f}" +
            (f" 🆕" if r["yeni_giris"] else
             f" ▲%{r['degisim']:.1f}" if r["degisim"] > 0.5 else
             f" ▼%{abs(r['degisim']):.1f}" if r["degisim"] < -0.5 else "")
            for _, r in grp.iterrows()
        ])

        toplam    = grp["oran2"].sum()
        kurum_say = len(grp)
        yeni_say  = grp["yeni_giris"].sum()
        artan_say = (grp["degisim"] > 0.5).sum()

        # Alarm seviyesi
        if kurum_say >= 3 and toplam >= 20:
            alarm = "🔴 KRİTİK"
        elif kurum_say >= 2 and toplam >= 10:
            alarm = "🟠 GÜÇLÜ"
        elif yeni_say > 0:
            alarm = "🟡 YENİ GİRİŞ"
        elif artan_say >= 1:
            alarm = "🟢 ARTIYOR"
        else:
            alarm = "⚪ İZLE"

        # Trend ikonu (son 3 ay)
        trend_str = ""
        if onceki_ay:
            trend_vals = []
            for d in donemler[-3:]:
                v = df_aylik[(df_aylik["hisse"]==hisse) & 
                             (df_aylik["donem"]==d) &
                             (df_aylik["kurum"].isin(KURUMLAR))]["oran2"].sum()
                trend_vals.append(v)
            if len(trend_vals) >= 2:
                if all(trend_vals[i] > trend_vals[i-1] for i in range(1, len(trend_vals))):
                    trend_str = "🚀 Sürekli Artış"
                elif trend_vals[-1] > trend_vals[-2]:
                    trend_str = "📈 Son Ay Arttı"
                elif trend_vals[-1] < trend_vals[-2]:
                    trend_str = "📉 Son Ay Azaldı"
                else:
                    trend_str = "➡️ Sabit"

        rows.append({
            "Hisse":        hisse,
            "Kurumlar":     kurumlar_str,
            "Toplam %":     round(toplam, 1),
            "Kurum Sayısı": kurum_say,
            "Trend":        trend_str,
            "Alarm":        alarm,
        })

    df_tablo = pd.DataFrame(rows)
    df_tablo = df_tablo[df_tablo["Kurum Sayısı"] >= int(min_kurum)]
    df_tablo = df_tablo.sort_values(["Alarm","Toplam %"], ascending=[True, False])

    if df_tablo.empty:
        st.warning("Filtre kriterlerine uyan hisse yok.")
        return

    # ── Özet metrikler ────────────────────────────────────────────────────────
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Toplam Hisse", len(df_tablo))
    c2.metric("🔴 Kritik",    len(df_tablo[df_tablo["Alarm"]=="🔴 KRİTİK"]))
    c3.metric("🟠 Güçlü",     len(df_tablo[df_tablo["Alarm"]=="🟠 GÜÇLÜ"]))
    c4.metric("🟡 Yeni Giriş",len(df_tablo[df_tablo["Alarm"]=="🟡 YENİ GİRİŞ"]))

    st.divider()

    # ── Alarm kartları ────────────────────────────────────────────────────────
    for _, r in df_tablo.iterrows():
        alarm = r["Alarm"]
        renk = {
            "🔴 KRİTİK":    "#C0392B",
            "🟠 GÜÇLÜ":     "#E67E22",
            "🟡 YENİ GİRİŞ":"#F39C12",
            "🟢 ARTIYOR":   "#1A7A3E",
            "⚪ İZLE":       "#717D7E",
        }.get(alarm, "#717D7E")

        st.markdown(
            f"""<div style='border-left:5px solid {renk};padding:8px 12px;
            margin:5px 0;background:#FAFAFA;border-radius:0 6px 6px 0;'>
            <span style='font-size:16px;font-weight:bold;'>{r['Hisse']}</span>
            <span style='color:{renk};margin-left:10px;font-weight:bold;'>{alarm}</span>
            <span style='float:right;color:#888;'>Toplam: <b>%{r['Toplam %']}</b> | {r['Kurum Sayısı']} kurum | {r['Trend']}</span><br>
            <small style='color:#555;'>{r['Kurumlar']}</small>
            </div>""",
            unsafe_allow_html=True
        )


def takas_sekmesi():
    """Ana takas analizi sekmesi."""

    # ── Üst Kontrol ──────────────────────────────────────────────────────────
    col_btn1, col_btn2, col_btn3, col_spacer = st.columns([1, 1, 1, 5])

    with col_btn1:
        mod_gunluk = st.button("📅 GÜNLÜK", use_container_width=True,
                               type="primary" if st.session_state.get("takas_mod") == "gunluk" else "secondary")
    with col_btn2:
        mod_haftalik = st.button("📆 HAFTALIK", use_container_width=True,
                                 type="primary" if st.session_state.get("takas_mod") == "haftalik" else "secondary")
    with col_btn3:
        mod_aylik = st.button("🗓️ AYLIK", use_container_width=True,
                              type="primary" if st.session_state.get("takas_mod") == "aylik" else "secondary")

    # Mod seçimi
    if mod_gunluk:
        st.session_state["takas_mod"] = "gunluk"
    elif mod_haftalik:
        st.session_state["takas_mod"] = "haftalik"
    elif mod_aylik:
        st.session_state["takas_mod"] = "aylik"

    if "takas_mod" not in st.session_state:
        st.session_state["takas_mod"] = "haftalik"

    mod = st.session_state["takas_mod"]

    st.markdown("---")

    # Dönem seçimi
    tip_map = {"gunluk": "gunluk", "haftalik": "haftalik", "aylik": "aylik"}
    donemler = donemler_listele(tip_map[mod])

    if not donemler:
        st.info(f"📂 {mod.upper()} veri yok. **Veri Yükle** sekmesinden ekleyin.")
        return

    col_d1, col_d2, col_d3 = st.columns([3, 1, 1])
    with col_d1:
        if mod == "gunluk":
            secili_donemler = st.multiselect(
                "Günler:", donemler, default=donemler[:3],
                key="takas_donem_sec"
            )
        elif mod == "haftalik":
            secili_donemler = st.multiselect(
                "Haftalar:", donemler, default=donemler[:4],
                key="takas_donem_sec"
            )
        else:
            secili_donemler = st.multiselect(
                "Aylar:", donemler, default=donemler[:3],
                key="takas_donem_sec"
            )
    with col_d2:
        min_pct = st.number_input("Min %:", value=0.5, step=0.1,
                                  min_value=0.0, max_value=10.0,
                                  key="takas_min_pct")
    with col_d3:
        st.markdown("<br>", unsafe_allow_html=True)
        ilk_giris = st.checkbox("🔵 İlk Giriş (%0→%3)", key="takas_ilk_giris")

    if not secili_donemler:
        st.warning("Dönem seçin.")
        return

    # ── Seçilen Dönemde %3+ Artış Yapan Kurumlar ────────────────────────────
    from takas_depo import _oku as takas_oku_raw
    t2_df = takas_oku_raw()

    if not t2_df.empty and secili_donemler:
        t2_sec = t2_df[t2_df["donem"].isin(secili_donemler)].copy()

        # Her hisse+kurum için seçilen dönemde toplam artış
        artis_df = t2_sec.groupby(["hisse", "kurum"]).agg(
            dolasim_pct=("dolasim_pct", "sum"),
            oran2=("oran2", "last")
        ).reset_index()

        if ilk_giris:
            # İlk giriş modu: %0'dan %3'e kadar alım yapanlar
            df_onceki = t2_df[~t2_df["donem"].isin(secili_donemler)].copy()
            onceki = df_onceki.groupby(["hisse", "kurum"])["oran2"].last().reset_index()
            onceki.columns = ["hisse", "kurum", "onceki_oran"]

            # tks2 bilgisini de ekle
            tks2_df = t2_sec.groupby("hisse")["tks2"].last().reset_index()
            artis_df = artis_df.merge(onceki, on=["hisse", "kurum"], how="left")
            artis_df = artis_df.merge(tks2_df, on="hisse", how="left")
            artis_df["onceki_oran"] = artis_df["onceki_oran"].fillna(0)
            artis_df["tks2"] = artis_df["tks2"].fillna(0)

            # Dinamik eşik — tahta büyüklüğüne göre
            def dinamik_esik(tks2):
                if tks2 >= 500_000_000:   # 500M+ lot → büyük tahta
                    return 0.3
                elif tks2 >= 100_000_000: # 100-500M lot → orta tahta
                    return 1.0
                else:                      # 100M altı → küçük tahta
                    return 2.0

            artis_df["esik"] = artis_df["tks2"].apply(dinamik_esik)

            artis_df = artis_df[
                (artis_df["onceki_oran"] < 0.5) &
                (artis_df["dolasim_pct"] >= artis_df["esik"]) &
                (artis_df["oran2"] <= 3)
            ].copy()
            baslik = "### 🔵 İlk Giriş Yapan Kurumlar (%0 → %3)"
        else:
            # Normal mod: %3+ artış yapanlar
            artis_df = artis_df[artis_df["dolasim_pct"] >= max(3, min_pct)].copy()
            baslik = "### 📈 Seçilen Dönemde %3+ Artış Yapan Kurumlar"

        if not artis_df.empty:
            st.markdown(baslik)
            kurumlar_listesi = sorted(artis_df["kurum"].unique())
            cols = st.columns(min(len(kurumlar_listesi), 4))

            for i, kurum in enumerate(kurumlar_listesi):
                col = cols[i % 4]
                k_df = artis_df[artis_df["kurum"] == kurum].sort_values("dolasim_pct", ascending=False).head(5)
                renk = "#1A5276" if kurum in AKILLI_PARA else "#1A7A3E"
                with col:
                    st.markdown(
                        "<div style='font-weight:bold;color:" + renk + ";margin-bottom:4px;'>"
                        "📈 " + kurum + "</div>",
                        unsafe_allow_html=True
                    )
                    for _, r in k_df.iterrows():
                        st.markdown(
                            "<div style='font-size:12px;padding:2px 0;'>"
                            "<b>" + r['hisse'] + "</b> "
                            "<span style='color:" + renk + ";font-weight:bold;'>+"
                            + f"{r['dolasim_pct']:.1f}%" + "</span>"
                            "</div>",
                            unsafe_allow_html=True
                        )

    st.markdown("---")

    # ── Alarm Tablosu ─────────────────────────────────────────────────────────
    col_alarm, col_trend = st.columns([1, 1])

    with col_alarm:
        st.markdown("### 🚨 Toplama Alarm Listesi")
        alarm_df = alarm_listesi(secili_donemler, min_pct)

        if False and not alarm_df.empty:  # ilk_giris ile değiştirildi
            alarm_df = alarm_df[alarm_df["akilli_para"] == True]

        if alarm_df.empty:
            st.caption("Bu dönemde %3+ artış yapan kurum yok.")
        else:
            # Hisse bazlı grupla
            hisse_grp = {}
            alarm_sira_map = {"🔴 KRİTİK": 0, "🟠 GÜÇLÜ": 1}
            for _, r in alarm_df.iterrows():
                h = r["hisse"]
                if h not in hisse_grp:
                    hisse_grp[h] = {"alarm": r["alarm"], "kurumlar": []}
                if alarm_sira_map.get(r["alarm"], 9) < alarm_sira_map.get(hisse_grp[h]["alarm"], 9):
                    hisse_grp[h]["alarm"] = r["alarm"]
                hisse_grp[h]["kurumlar"].append(r)

            for hisse, data in list(hisse_grp.items())[:30]:
                alarm = data["alarm"]
                renk = {"🔴 KRİTİK": "#C0392B", "🟠 GÜÇLÜ": "#E67E22"}.get(alarm, "#95A5A6")

                kurum_html = ""
                for r in sorted(data["kurumlar"], key=lambda x: -x["dolasim_pct"]):
                    kurum_html += (
                        "<div style='font-size:12px;padding:1px 0 1px 8px;'>"
                        f"<b style='color:{renk};'>{r['kurum']}</b>"
                        f" &nbsp;|&nbsp; T2: %{r['oran2']:.2f}"
                        f" &nbsp;|&nbsp; Artış: <b>+{r['dolasim_pct']:.2f}%</b>"
                        "</div>"
                    )

                st.markdown(
                    "<div style='border-left:4px solid " + renk + ";padding:6px 10px;"
                    "margin:4px 0;background:#FAFAFA;border-radius:0 4px 4px 0;'>"
                    "<b style='font-size:14px;'>" + hisse + "</b>"
                    "<span style='color:" + renk + ";margin-left:8px;font-weight:bold;'>" + alarm + "</span>"
                    + kurum_html +
                    "</div>",
                    unsafe_allow_html=True
                )

    with col_trend:
        st.markdown("### 📊 Trend Analizi")
        trend_df = trend_analiz(min_hafta=2)

        if False and not trend_df.empty:  # ilk_giris ile değiştirildi
            trend_df = trend_df[trend_df["kurum"].isin(AKILLI_PARA)]

        if trend_df.empty:
            st.caption("Trend verisi yok (min 2 dönem gerekli).")
        else:
            for _, r in trend_df.head(15).iterrows():
                trend = r["trend"]
                renk = "#1A5276" if "Sürekli" in trend else "#1A7A3E"
                st.markdown(
                    f"""<div style='border-left:4px solid {renk};padding:6px 10px;
                         margin:4px 0;background:#FAFAFA;border-radius:0 4px 4px 0;'>
                    <b>{r['hisse']}</b> — <span style='color:#888;'>{r['kurum']}</span>
                    <span style='color:{renk};margin-left:8px;'>{trend}</span><br>
                    <span style='font-size:12px;color:#555;'>
                    Son: +{r['son_pct']:.2f}% | Toplam: +{r['toplam_pct']:.2f}% | {r['hafta_sayisi']} dönem
                    </span></div>""",
                    unsafe_allow_html=True
                )

    st.markdown("---")

    # ── Ana Tablo ─────────────────────────────────────────────────────────────
    st.markdown("### 📋 Kurum Bazlı Detay Tablo")

    pivot_df = takas_analiz(secili_donemler, tip_map[mod])

    if not pivot_df.empty:
        kurum_cols = [c for c in pivot_df.columns
                      if c not in ["hisse", "tks2"]]

        # Toplam kolon
        pivot_df["TOPLAM"] = pivot_df[kurum_cols].sum(axis=1).round(2)

        # Filtre
        if False:  # ilk_giris ile değiştirildi
            akilli_cols = [c for c in kurum_cols if c in AKILLI_PARA]
            if akilli_cols:
                pivot_df = pivot_df[pivot_df[akilli_cols].max(axis=1) >= min_pct]
        else:
            pivot_df = pivot_df[pivot_df["TOPLAM"].abs() >= min_pct]

        pivot_df = pivot_df.sort_values("TOPLAM", ascending=False)

        if pivot_df.empty:
            st.caption(f"Min %{min_pct} eşiğini geçen hisse yok.")
        else:
            # Renklendirme
            fmt = {c: "{:+.2f}" for c in kurum_cols + ["TOPLAM"]}
            fmt["tks2"] = "{:,.0f}"

            def renk_pct(val):
                if isinstance(val, float):
                    if val >= 1.0:
                        return "background-color:#D5F5E3;color:#1A5276;font-weight:bold"
                    if val > 0:
                        return "color:#1A5276;font-weight:bold"
                    if val <= -1.0:
                        return "background-color:#FADBD8;color:#C0392B;font-weight:bold"
                    if val < 0:
                        return "color:#C0392B"
                return "color:#888888"

            styled = pivot_df.style.map(renk_pct, subset=kurum_cols + ["TOPLAM"])
            styled = styled.format(fmt, na_rep="—")

            st.dataframe(styled, use_container_width=True, height=500)

            # Excel indir
            buf = _takas_excel_indir(pivot_df, secili_donemler)
            st.download_button(
                "⬇️ Excel İndir",
                data=buf,
                file_name=f"takas_analiz_{mod}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


def takas_veri_yukle_bolumu():
    """Veri Yükle sekmesindeki takas bölümü."""
    st.markdown("### 🏦 Kurum Takas Dosyaları")
    st.caption(
        "Dosya adı formatı: **KURUM_AY_DONEM.xlsx**  |  "
        "Aylık: `TERA__2026_01_03.xlsx`  |  "
        "Haftalık: `TERA_202604_01.xlsx`  |  "
        "Günlük: `TERA_20260420.xlsx`"
    )

    with st.form("takas_yukle_form"):
        yuklenen = st.file_uploader(
            "📂 Dosyaları Seçin (birden fazla seçilebilir):",
            type=["xlsx"],
            accept_multiple_files=True,
            key="takas_dosya_yukle"
        )
        yukle_btn = st.form_submit_button(
            "✅ Yükle", use_container_width=True
        )

    if yukle_btn:
        if not yuklenen:
            st.error("En az 1 dosya seçin!")
        else:
            dosya_listesi = [(f.name, f) for f in yuklenen]
            eklenen, hatalar = dosyalar_yukle(dosya_listesi)

            if eklenen:
                for msg in eklenen:
                    st.success(msg)
            if hatalar:
                for msg in hatalar:
                    st.warning(msg)

            if eklenen:
                st.rerun()

    # Kayıtlı dönemler
    st.markdown("---")
    st.markdown("**📂 Kayıtlı Veriler:**")

    col_g, col_h, col_a = st.columns(3)

    for col, tip, label in [
        (col_g, "gunluk", "📅 Günlük"),
        (col_h, "haftalik", "📆 Haftalık"),
        (col_a, "aylik", "🗓️ Aylık")
    ]:
        with col:
            st.markdown(f"**{label}:**")
            donemler = donemler_listele(tip)
            if donemler:
                kurumlar = kurumlar_listele()
                for d in donemler[:10]:
                    c1, c2 = st.columns([3, 1])
                    c1.markdown(f"`{d}`")
                    if c2.button("🗑️", key=f"takas_sil_{tip}_{d}"):
                        for k in kurumlar:
                            donem_sil(k, d)
                        st.rerun()
            else:
                st.caption("Henüz yok")


def _takas_excel_indir(df: pd.DataFrame, donemler: list):
    """Takas tablosunu Excel'e dönüştürür."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Takas Analiz"

    cols = list(df.columns)
    tarih = datetime.now().strftime("%d.%m.%Y")

    # Başlık
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    c = ws["A1"]
    c.value = f"Takas Analizi | {', '.join(donemler)} | {tarih}"
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="1A252F")
    c.alignment = Alignment(horizontal="center")

    # Header
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=2, column=i, value=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", start_color="1A3A5C")
        cell.alignment = Alignment(horizontal="center")

    # Veri
    for ri, (_, row) in enumerate(df.iterrows()):
        r = ri + 3
        for ci, col in enumerate(cols, 1):
            v = row[col]
            cell = ws.cell(row=r, column=ci, value=v)
            if isinstance(v, float):
                cell.number_format = "+0.00;-0.00;0.00"
                if v >= 1.0:
                    cell.font = Font(bold=True, color="1A5276")
                elif v > 0:
                    cell.font = Font(color="1A5276")
                elif v <= -1.0:
                    cell.font = Font(bold=True, color="C0392B")
                elif v < 0:
                    cell.font = Font(color="C0392B")

    # Kolon genişlikleri
    ws.column_dimensions["A"].width = 10
    for i in range(2, len(cols) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.freeze_panes = "B3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
